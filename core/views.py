from django.utils.translation import gettext as _
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Avg, Q, Case, When, IntegerField
from django.utils import timezone
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import datetime, timedelta, date as _date
from django.db import models
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from .models import Direction, Project, Document, Partner, Event, EventMember, Request, Employee, Budget, UserProfile, UserActivity, ProjectMember, Milestone, SubMilestone, ProjectNeed, ProjectComment, ProjectDocument, ProjectFolder, ProjectActivity, Role, ProjectRole, LeaveRequest
from .notifs import push_section_refresh, push_project_meta, push_milestone_status


def log_project_activity(project, action, description, user):
    """Enregistre une activité sur un projet"""
    username = user.get_full_name() or user.username if hasattr(user, 'get_full_name') else str(user)
    ProjectActivity.objects.create(
        project=project,
        action=action,
        description=description,
        user=username
    )
    push_section_refresh(project.id, 'panel-activite', username)


def _ensure_manager_in_team(project, user=None):
    """S'assure que le responsable (manager_employee) fait partie de l'équipe projet."""
    manager = project.manager_employee
    if not manager:
        return None

    manager_role, __ = ProjectRole.objects.get_or_create(
        slug='responsable',
        defaults={
            'name': _('Responsable'),
            'description': _('Responsable du projet'),
            'is_system': True,
        }
    )

    member, created = ProjectMember.objects.get_or_create(
        project=project,
        employee=manager,
        defaults={'project_role': manager_role}
    )

    if not created and member.project_role_id != manager_role.id:
        member.project_role = manager_role
        member.save(update_fields=['project_role'])

    if created and user:
        log_project_activity(
            project, 'ajout_membre',
            _("Ajout automatique du responsable '%(name)s' à l'équipe") % {'name': manager.name},
            user
        )

    return member


_AVATAR_COLORS = [
    '#2a4a6f',  # --primary-500
    '#16a34a',  # --green-600
    '#d97706',  # --amber-600
    '#dc2626',  # --red-600
    '#9333ea',  # --purple-600
    '#1c3550',  # --primary-700
    '#15803d',  # --green-700
    '#b45309',  # --amber-700
    '#b91c1c',  # --red-700
    '#7c3aed',  # --purple-700
]


def _get_project_member_cards(project):
    """Return (project_members, other_employees) as lists of card dicts for the assignment picker."""
    _project_members = list(
        project.members.filter(employee__isnull=False).select_related('project_role')
    )
    member_pks = {pm.employee_id for pm in _project_members}
    pm_roles = {
        pm.employee_id: pm.project_role.name if pm.project_role else '—'
        for pm in _project_members
    }
    m_counts = dict(
        Employee.objects.filter(assigned_milestones__project=project)
        .values('id').annotate(n=Count('assigned_milestones', distinct=True))
        .values_list('id', 'n')
    )
    s_counts = dict(
        Employee.objects.filter(assigned_sub_milestones__milestone__project=project)
        .values('id').annotate(n=Count('assigned_sub_milestones', distinct=True))
        .values_list('id', 'n')
    )
    all_emps = list(Employee.objects.select_related('direction').order_by('name'))
    members, others = [], []
    for emp in all_emps:
        parts = emp.name.split()
        initials = parts[0][0].upper() if parts else '?'
        if len(parts) > 1:
            initials += parts[-1][0].upper()
        card = {
            'id': emp.id,
            'name': emp.name,
            'job_role': emp.role,
            'project_role': pm_roles.get(emp.id, ''),
            'direction': emp.direction.code if emp.direction else (emp.organization or ''),
            'task_count': m_counts.get(emp.id, 0) + s_counts.get(emp.id, 0),
            'initials': initials,
            'color': _AVATAR_COLORS[sum(ord(c) for c in emp.name) % len(_AVATAR_COLORS)],
            'is_external': emp.is_external,
        }
        if emp.id in member_pks:
            members.append(card)
        else:
            others.append(card)
    return members, others


def _user_is_assignee(user, assigned_m2m):
    """Vérifie si l'utilisateur connecté est parmi les responsables d'une tâche (M2M)."""
    profile_employee = getattr(getattr(user, 'profile', None), 'employee', None)
    if not profile_employee:
        return False
    try:
        return assigned_m2m.filter(pk=profile_employee.pk).exists()
    except (AttributeError, ValueError) as e:
        from .error_logging import ErrorLogger
        ErrorLogger.log_exception(e, context={'function': '_user_is_assignee'}, user=user)
        return False


def get_accessible_projects_qs(user):
    """Retourne le queryset de base des projets accessibles selon le rôle de l'utilisateur."""
    from django.db.models import Q
    profile = user.profile
    qs = Project.objects.select_related('direction')
    if profile.is_directeur_general() or profile.is_admin():
        return qs
    # RBAC : permission globale de lecture sur tous les projets
    if profile.can_view_all_projects():
        return qs

    # Permission conditionnelle (ex : same_direction) → scope à la direction
    if profile.can('read', 'Project') or profile.can('manage', 'Project'):
        if profile.direction_id:
            return qs.filter(direction_id=profile.direction_id)
        return qs.none()

    employee_id = getattr(profile, 'employee_id', None)

    # External users: only see projects where they are a ProjectMember
    if employee_id:
        try:
            emp = profile.employee
            if emp and emp.is_external:
                return qs.filter(members__employee_id=employee_id).distinct()
        except (AttributeError, ValueError) as e:
            from .error_logging import ErrorLogger
            ErrorLogger.log_exception(e, context={'function': 'get_accessible_projects_qs'}, user=user)

    # Sans lien employé, aucun accès par appartenance
    if not employee_id:
        return qs.none()

    member_q = Q(members__employee_id=employee_id)
    manager_q = Q(manager_employee_id=employee_id)
    return qs.filter(manager_q | member_q).distinct()


# ── Helpers comptes utilisateurs ─────────────────────────────────────────────

def _generate_username(full_name):
    """Génère un username unique au format nom.prenom (sans accents, ASCII uniquement)."""
    import unicodedata
    import re
    from django.contrib.auth.models import User as _User

    def _norm(s):
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return re.sub(r'[^a-z0-9]', '', s.lower())

    parts = full_name.strip().split()
    if len(parts) >= 2:
        base = f"{_norm(parts[-1])}.{_norm(parts[0])}"
    elif parts:
        base = _norm(parts[0])
    else:
        base = 'utilisateur'

    username, counter = base, 2
    while _User.objects.filter(username=username).exists():
        username = f"{base}{counter}"
        counter += 1
    return username


def _create_or_update_user_for_employee(request, employee, system_role):
    """Crée ou met à jour le compte utilisateur lié à un employé et envoie l'invitation."""
    from django.contrib.auth.models import User as DjangoUser
    from .models import UserProfile

    if not employee.email:
        return False, _("L'employé n'a pas d'adresse email.")

    try:
        linked_profile = employee.user_profile
        linked_profile.role = system_role
        linked_profile.direction = employee.direction
        linked_profile.save(update_fields=['role', 'direction', 'updated_at'])
        return True, _("Le rôle système du compte lié a été mis à jour.")
    except (AttributeError, ValueError) as e:
        from .error_logging import ErrorLogger
        ErrorLogger.log_exception(e, context={'function': '_create_or_update_user_for_employee', 'employee': employee.name}, user=request.user)

    username = _generate_username(employee.name)
    name_parts = employee.name.strip().split()

    user = DjangoUser(
        username=username,
        email=employee.email,
        first_name=name_parts[0] if name_parts else '',
        last_name=' '.join(name_parts[1:]) if len(name_parts) > 1 else '',
        is_active=False,
    )
    user.set_unusable_password()
    user.save()

    profile, created = UserProfile.objects.get_or_create(user=user)
    profile.role = system_role
    profile.direction = employee.direction
    profile.employee = employee
    profile.save()

    activation_link = _build_activation_link(request, user)
    invited_by = request.user.get_full_name() or request.user.username
    from .notifications import notify_account_invitation
    email_ok, email_msg = notify_account_invitation(user, activation_link, invited_by)

    UserActivity.objects.create(
        user=request.user,
        action='create',
        description=f"Création du compte {username} pour {employee.name}",
        ip_address=get_client_ip(request),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
    )

    if email_ok:
        return True, _("Compte créé pour {name}. Invitation envoyée à {email}.").format(name=employee.name, email=employee.email)
    return True, _("Compte créé (identifiant : {username}) mais l'email n'a pas pu être envoyé : {msg}").format(username=username, msg=email_msg)


def _build_reset_link(request, user):
    """Génère le lien de réinitialisation de mot de passe (uid + token Django, valable 3 jours)."""
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_encode
    from django.utils.encoding import force_bytes
    from django.urls import reverse

    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    return request.build_absolute_uri(
        reverse('core:password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
    )


def _build_activation_link(request, user):
    """Génère le lien d'activation sécurisé (uid + token Django, valable 3 jours)."""
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_encode
    from django.utils.encoding import force_bytes
    from django.urls import reverse

    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    return request.build_absolute_uri(
        reverse('core:account_activate', kwargs={'uidb64': uid, 'token': token})
    )


def login_view(request):
    """Vue de connexion"""
    if request.user.is_authenticated:
        return redirect('core:dashboard')

    if request.method == 'GET':
        storage = messages.get_messages(request)
        list(storage)

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                UserActivity.objects.create(
                    user=user, action='login',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                )
                from django.utils.http import url_has_allowed_host_and_scheme
                next_url = request.GET.get('next')
                if next_url and url_has_allowed_host_and_scheme(
                    next_url,
                    allowed_hosts={request.get_host()},
                    require_https=request.is_secure(),
                ):
                    return redirect(next_url)
                return redirect('core:dashboard')
            elif not user.has_usable_password():
                messages.error(request, _("Votre compte est en attente d'activation. Consultez votre email pour le lien d'invitation."))
            else:
                messages.error(request, _('Votre compte a été désactivé. Contactez un administrateur.'))
        else:
            messages.error(request, _('Identifiants incorrects.'))
    
    return render(request, 'registration/login.html')


@login_required
def logout_view(request):
    """Vue de déconnexion"""
    if request.method == 'POST':
        if request.user.is_authenticated:
            UserActivity.objects.create(
                user=request.user, action='logout',
                ip_address=request.META.get('REMOTE_ADDR'),
            )
        logout(request)
        return redirect(reverse('core:login') + '?logout=1')
    return redirect('core:dashboard')


def password_reset_request(request):
    """Formulaire de demande de réinitialisation de mot de passe (publique)."""
    from django.contrib.auth.models import User as DjangoUser

    sent = False
    form_error = None

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            form_error = "Veuillez saisir votre adresse email."
        else:
            try:
                user = DjangoUser.objects.get(email__iexact=email, is_active=True)
            except DjangoUser.DoesNotExist:
                user = None

            if user is not None and user.has_usable_password():
                reset_link = _build_reset_link(request, user)
                from .notifications import notify_password_reset
                notify_password_reset(user, reset_link)

            # Réponse neutre : ne pas révéler si l'email existe
            sent = True

    return render(request, 'registration/password_reset.html', {
        'sent': sent,
        'form_error': form_error,
        'email_value': request.POST.get('email', '') if not sent else '',
    })


def password_reset_confirm(request, uidb64, token):
    """Confirmation de réinitialisation — saisie du nouveau mot de passe."""
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    from django.utils.encoding import force_str
    from django.contrib.auth.models import User as DjangoUser
    from django.contrib.auth.forms import SetPasswordForm

    user = None
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = DjangoUser.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, DjangoUser.DoesNotExist):
        pass

    token_valid = (
        user is not None
        and user.is_active
        and user.has_usable_password()
        and default_token_generator.check_token(user, token)
    )

    if not token_valid:
        if user is not None and not user.has_usable_password():
            # Compte en attente d'activation — rediriger vers le bon flow
            return render(request, 'registration/password_reset_confirm.html', {
                'valid': False,
                'pending_account': True,
            })
        return render(request, 'registration/password_reset_confirm.html', {'valid': False})

    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)
        if form.is_valid():
            form.save()
            UserActivity.objects.create(
                user=user,
                action='update',
                description="Réinitialisation du mot de passe via email",
            )
            messages.success(request, _("Mot de passe réinitialisé. Vous pouvez vous connecter."))
            return redirect('core:login')
    else:
        form = SetPasswordForm(user)

    return render(request, 'registration/password_reset_confirm.html', {
        'valid': True,
        'form': form,
        'username': user.username,
        'full_name': user.get_full_name(),
    })


@login_required
def password_change(request):
    """L'utilisateur connecté change son propre mot de passe."""
    from django.contrib.auth.forms import PasswordChangeForm
    from django.contrib.auth import update_session_auth_hash

    form = PasswordChangeForm(request.user, request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        update_session_auth_hash(request, user)
        UserActivity.objects.create(
            user=request.user,
            action='update',
            description="Changement de mot de passe",
            ip_address=get_client_ip(request),
        )
        messages.success(request, _("Mot de passe modifié avec succès."))
        return redirect('core:dashboard')

    return render(request, 'registration/password_change.html', {'form': form})


@login_required
def profile(request):
    """Page de profil : informations personnelles de l'utilisateur connecté."""
    user = request.user
    prof = user.profile
    errors = {}

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        phone      = request.POST.get('phone', '').strip()
        avatar     = request.POST.get('avatar', '').strip()

        if request.POST.get('remove_avatar'):
            prof.avatar = ''
            prof.save(update_fields=['avatar', 'updated_at'])
            messages.success(request, _("Photo de profil supprimée."))
            return redirect('core:profile')

        if not first_name:
            errors['first_name'] = "Le prénom est requis."
        if not last_name:
            errors['last_name'] = "Le nom est requis."
        if not email:
            errors['email'] = "L'adresse email est requise."
        elif email != user.email:
            from django.contrib.auth.models import User as DjangoUser
            if DjangoUser.objects.filter(email__iexact=email).exclude(pk=user.pk).exists():
                errors['email'] = "Cette adresse email est déjà utilisée par un autre compte."

        # Validation du téléphone (format international)
        normalized_phone = phone
        if phone:
            from core.fields import validate_phone_number
            try:
                normalized_phone = validate_phone_number(phone)
            except (ValueError, AttributeError) as exc:
                from .error_logging import ErrorLogger
                ErrorLogger.log_validation_error('phone', str(exc), user=user)
                errors['phone'] = str(exc)

        if not errors:
            user.first_name = first_name
            user.last_name  = last_name
            user.email      = email
            user.save(update_fields=['first_name', 'last_name', 'email'])
            prof.phone = normalized_phone
            prof.avatar = avatar
            prof.save(update_fields=['phone', 'avatar', 'updated_at'])
            UserActivity.objects.create(
                user=user,
                action='update',
                description="Mise à jour du profil",
                ip_address=get_client_ip(request),
            )
            messages.success(request, _("Profil mis à jour avec succès."))
            return redirect('core:profile')

        return render(request, 'core/profile.html', {
            'form_data': request.POST,
            'errors': errors,
        })

    return render(request, 'core/profile.html', {})


@login_required
def global_search(request):
    """Recherche globale depuis la barre du header"""
    query = request.GET.get('q', '').strip()
    results = {
        'projects': [],
        'partners': [],
        'employees': [],
        'requests': [],
        'documents': [],
    }
    
    if query:
        from django.db.models import Q
        _profile = request.user.profile
        _dir_id = getattr(_profile, 'direction_id', None)
        _is_global = _profile.is_admin() or _profile.is_directeur_general()

        # Projets — scopés via RBAC
        _accessible = get_accessible_projects_qs(request.user)
        results['projects'] = _accessible.select_related('direction').filter(
            Q(name__icontains=query) | Q(manager__icontains=query) | Q(description__icontains=query)
        )[:10]

        # Partenaires — pas de direction, visibles si can_read_partners
        if _profile.can_read_partners():
            results['partners'] = Partner.objects.filter(
                Q(name__icontains=query) | Q(contact_person__icontains=query)
            )[:10]

        # Employés — scopés à la direction pour les non-globaux
        _emp_qs = Employee.objects.select_related('direction').filter(
            Q(name__icontains=query) | Q(role__icontains=query) | Q(email__icontains=query)
        )
        if not _is_global and _dir_id:
            _emp_qs = _emp_qs.filter(direction_id=_dir_id)
        elif not _is_global:
            _emp_qs = _emp_qs.none()
        results['employees'] = _emp_qs[:10]

        # Demandes — scopées à la direction
        _req_qs = Request.objects.select_related('direction').filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )
        if not _is_global:
            if _dir_id:
                _req_qs = _req_qs.filter(direction_id=_dir_id)
            else:
                _req_qs = _req_qs.none()
        results['requests'] = _req_qs[:10]

        # Documents — scopés à la direction
        _doc_qs = Document.objects.select_related('direction').filter(Q(title__icontains=query))
        if not _is_global:
            if _dir_id:
                _doc_qs = _doc_qs.filter(direction_id=_dir_id)
            else:
                _doc_qs = _doc_qs.none()
        results['documents'] = _doc_qs[:10]
    
    total = sum(len(v) for v in results.values())
    
    context = {
        'query': query,
        'results': results,
        'total': total,
    }
    return render(request, 'core/global_search.html', context)


def _project_budget_data(project_qs):
    """Calcule les données budget par projet pour un queryset de projets donné."""
    budget_by_project = {}
    project_ids = set(project_qs.values_list('id', flat=True))

    # Source 1 : champ budget directement sur le projet
    for p in project_qs.filter(Q(budget__gt=0) | Q(budget_consumed__gt=0)):
        budget_by_project[p.id] = {
            'name': p.name,
            'allocated': float(p.budget or 0),
            'consumed': float(p.budget_consumed or 0),
        }

    # Source 2 : lignes budgétaires rattachées à un projet (somme)
    project_budget_rows = (
        Budget.objects
        .filter(project_id__in=project_ids)
        .values('project_id', 'project__name')
        .annotate(total_allocated=Sum('allocated'), total_consumed=Sum('consumed'))
    )
    for row in project_budget_rows:
        pid = row['project_id']
        entry = budget_by_project.setdefault(pid, {
            'name': row['project__name'] or 'Projet sans nom',
            'allocated': 0.0,
            'consumed': 0.0,
        })
        entry['allocated'] += float(row['total_allocated'] or 0)
        entry['consumed'] += float(row['total_consumed'] or 0)

    sorted_entries = sorted(budget_by_project.values(), key=lambda e: e['allocated'], reverse=True)[:10]
    budget_data = []
    for entry in sorted_entries:
        name = entry['name']
        budget_data.append({
            'direction': name[:30] + ('…' if len(name) > 30 else ''),
            'name': name,
            'color': '#3b82f6',
            'allocated': entry['allocated'] / 1000000,
            'consumed': entry['consumed'] / 1000000,
        })
    return budget_data


def _project_stats(project_qs):
    """Retourne les compteurs de projets pour un queryset (1 seule requête)."""
    return project_qs.aggregate(
        total_projects=Count('id', distinct=True),
        projects_in_progress=Count(Case(When(status='en_cours', then=1), output_field=IntegerField())),
        projects_completed=Count(Case(When(status='termine', then=1), output_field=IntegerField())),
        projects_planned=Count(Case(When(status='planifie', then=1), output_field=IntegerField())),
    )


def _project_budget_totals(project_qs):
    """Calcule budget total alloué / consommé pour un queryset de projets."""
    project_ids = set(project_qs.values_list('id', flat=True))
    budget_rows = Budget.objects.filter(project_id__in=project_ids).aggregate(
        allocated=Sum('allocated'), consumed=Sum('consumed')
    )
    budget_table_allocated = budget_rows['allocated'] or 0
    budget_table_consumed = budget_rows['consumed'] or 0

    project_ids_with_budget_row = set(
        Budget.objects.filter(project_id__in=project_ids).values_list('project_id', flat=True)
    )
    project_direct = project_qs.exclude(id__in=project_ids_with_budget_row).aggregate(
        total_allocated=Sum('budget'), total_consumed=Sum('budget_consumed'),
    )
    project_direct_allocated = project_direct['total_allocated'] or 0
    project_direct_consumed = project_direct['total_consumed'] or 0

    total_budget = float(budget_table_allocated) + float(project_direct_allocated)
    total_consumed = float(budget_table_consumed) + float(project_direct_consumed)
    budget_percentage = round((total_consumed / total_budget * 100), 1) if total_budget > 0 else 0
    return total_budget, total_consumed, budget_percentage


def _dashboard_scope_qs(user, profile, accessible_projects):
    """Retourne les querysets de documents, demandes, partenaires, événements et employés
    visibles depuis le dashboard selon le rôle/direction de l'utilisateur."""
    today = timezone.now().date()
    direction_id = getattr(profile, 'direction_id', None)
    is_global = profile and (profile.is_admin() or profile.is_directeur_general()) or user.is_superuser

    if is_global:
        docs_qs = Document.objects.exclude(status='signe')
        reqs_qs = Request.objects.filter(status='en_attente')
        partners_qs = Partner.objects.filter(status='actif') if profile.can_read_partners() else Partner.objects.none()
        events_qs = Event.objects.filter(date__gte=today)
        employees_qs = Employee.objects.all()
    elif direction_id:
        docs_qs = Document.objects.exclude(status='signe').filter(direction_id=direction_id)
        reqs_qs = Request.objects.filter(status='en_attente', direction_id=direction_id) if (profile and profile.can_approve_requests()) else Request.objects.none()
        partners_qs = Partner.objects.filter(status='actif') if (profile and profile.can_read_partners()) else Partner.objects.none()
        events_qs = Event.objects.filter(date__gte=today, participants__id=direction_id)
        employees_qs = Employee.objects.filter(direction_id=direction_id)
    else:
        docs_qs = Document.objects.none()
        reqs_qs = Request.objects.none()
        partners_qs = Partner.objects.none()
        events_qs = Event.objects.none()
        employees_qs = Employee.objects.none()

    return docs_qs, reqs_qs, partners_qs, events_qs, employees_qs


def _dashboard_pending_leaves(user, profile, direction_id):
    """Retourne les demandes de congé en attente visibles par l'utilisateur."""
    if not profile:
        return LeaveRequest.objects.none()

    # Admin / DG / rôle avec approbation finale
    if user.is_superuser or profile.can_give_final_approval():
        return LeaveRequest.objects.filter(
            status__in=['rh_conforme', 'avis_favorable']
        ).select_related('employee', 'direction').order_by('-created_at')[:5]

    # Rôles avec avis hiérarchique (directeur ou custom)
    if profile.can_give_manager_approval():
        qs = LeaveRequest.objects.filter(status='soumise')
        if direction_id:
            qs = qs.filter(direction_id=direction_id)
        return qs.select_related('employee', 'direction').order_by('-created_at')[:5]

    return LeaveRequest.objects.none()


# ==================== NOTIFICATIONS IN-APP ====================

@login_required
def notifications_list(request):
    """Page complète des notifications de l'utilisateur connecté."""
    notifs = request.user.notifications_received.all()[:50]
    request.user.notifications_received.filter(is_read=False).update(is_read=True)
    return render(request, 'core/notifications_list.html', {'notifications': notifs})


@login_required
def notifications_count(request):
    """API JSON : nombre de notifications non lues."""
    from django.http import JsonResponse
    count = request.user.notifications_received.filter(is_read=False).count()
    return JsonResponse({'unread': count})


@login_required
def notifications_recent(request):
    """API JSON : 6 dernières notifications avec détail."""
    from django.http import JsonResponse
    qs = request.user.notifications_received.all()[:6]
    data = [
        {
            'id': n.id,
            'type': n.notif_type,
            'title': n.title,
            'message': n.message,
            'link': n.link,
            'is_read': n.is_read,
            'created_at': n.created_at.strftime('%d/%m à %H:%M'),
        }
        for n in qs
    ]
    unread = request.user.notifications_received.filter(is_read=False).count()
    return JsonResponse({'notifications': data, 'unread': unread})


@login_required
def notification_mark_read(request, notif_id):
    """Marquer une notification comme lue (POST)."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    try:
        n = request.user.notifications_received.get(pk=notif_id)
        n.mark_read()
    except (AttributeError, ValueError) as e:
        from .error_logging import ErrorLogger
        ErrorLogger.log_exception(e, context={'function': 'notification_mark_read', 'notif_id': notif_id}, user=request.user)
    return JsonResponse({'ok': True})


@login_required
def notifications_mark_all_read(request):
    """Marquer toutes les notifications comme lues (POST)."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'method not allowed'}, status=405)
    request.user.notifications_received.filter(is_read=False).update(is_read=True)
    return JsonResponse({'ok': True})


# ==================== DASHBOARD ====================

@login_required
def dashboard(request):
    """Vue principale du tableau de bord, personnalisée selon le rôle."""
    today = timezone.now().date()
    user = request.user
    profile = getattr(user, 'profile', None)
    role_slug = profile.role_slug if profile else None
    employee_id = getattr(profile, 'employee_id', None)
    direction_id = getattr(profile, 'direction_id', None)

    accessible_projects = get_accessible_projects_qs(user)
    project_stats = _project_stats(accessible_projects)
    total_budget, total_consumed, budget_percentage = _project_budget_totals(accessible_projects)
    budget_data = _project_budget_data(accessible_projects)

    # Documents et demandes filtrés par scope accessible
    docs_qs, reqs_qs, partners_qs, events_qs, employees_qs = _dashboard_scope_qs(
        user, profile, accessible_projects
    )

    active_projects = accessible_projects.filter(status='en_cours')[:3]
    pending_docs = docs_qs.select_related('direction')[:4]
    pending_reqs = reqs_qs.select_related('direction')[:4]
    upcoming_events = events_qs.prefetch_related('participants')[:4]
    active_partners = partners_qs.count()

    # Données spécifiques selon le rôle
    role_context = {}

    _is_global_dash = user.is_superuser or (profile and (profile.is_admin() or profile.is_directeur_general()))
    _is_director_dash = profile and profile.can_view_all_projects() and direction_id and not _is_global_dash

    if _is_global_dash:
        role_context['dashboard_title'] = _("Vue d'ensemble de la CSIG")
        role_context['show_global_kpis'] = True

    elif _is_director_dash:
        role_context['dashboard_title'] = _("Vue de votre direction")
        role_context['show_global_kpis'] = False
        role_context['direction'] = profile.direction if profile else None
        if employees_qs.exists():
            avg_workload = employees_qs.aggregate(avg=Avg('workload'))['avg'] or 0
            role_context['avg_workload'] = round(avg_workload, 1)
            role_context['direction_employees_count'] = employees_qs.count()

    else:
        # Employé / chef de projet
        role_context['dashboard_title'] = _("Votre espace de travail")
        role_context['show_global_kpis'] = False

        # Mes projets
        role_context['my_projects_count'] = accessible_projects.count()

        # Mes tâches (jalons et sous-jalons assignés)
        if employee_id:
            my_milestones = Milestone.objects.filter(
                project__in=accessible_projects,
                assigned_to__id=employee_id
            ).exclude(status='termine').select_related('project').order_by('due_date')[:6]
            my_submilestones = SubMilestone.objects.filter(
                milestone__project__in=accessible_projects,
                assigned_to__id=employee_id
            ).exclude(completed=True).select_related('milestone', 'milestone__project').order_by('due_date')[:6]
            role_context['my_milestones'] = my_milestones
            role_context['my_submilestones'] = my_submilestones
            role_context['my_tasks_count'] = my_milestones.count() + my_submilestones.count()

            # Mes demandes de congé
            role_context['my_leave_requests'] = LeaveRequest.objects.filter(
                employee_id=employee_id
            ).order_by('-created_at')[:5]
        else:
            role_context['my_milestones'] = Milestone.objects.none()
            role_context['my_submilestones'] = SubMilestone.objects.none()
            role_context['my_tasks_count'] = 0
            role_context['my_leave_requests'] = LeaveRequest.objects.none()

    # Alertes projets en retard (visible si scope global sur les projets)
    if user.is_superuser or (profile and (profile.is_admin() or profile.is_directeur_general() or profile.can_view_all_projects())):
        role_context['late_projects'] = accessible_projects.filter(
            ~Q(status='termine'), end_date__lt=today
        ).order_by('end_date')[:5]

    # Congés en attente (selon les permissions)
    role_context['pending_leaves'] = _dashboard_pending_leaves(user, profile, direction_id)

    # Permissions RBAC pour affichage conditionnel des widgets
    rbac_context = {}
    if profile:
        rbac_context.update({
            'can_view_budgets': profile.can_view_budgets(),
            'can_read_partners': profile.can_read_partners(),
            'can_approve_requests': profile.can_approve_requests(),
            'can_give_final_approval': profile.can_give_final_approval(),
            'can_give_hr_check': profile.can_give_hr_check(),
            'can_view_reports': profile.can_view_reports(),
            'can_manage_users': profile.can_manage_users(),
            'can_manage_events': profile.can_manage_events(),
            'can_approve_documents': profile.can_approve_documents(),
            'can_view_all_projects': profile.can_view_all_projects(),
        })

    context = {
        'today': today,
        'role_slug': role_slug,
        'is_admin_or_dg': role_slug in ('admin', 'directeur_general') or user.is_superuser,
        'is_director': role_slug == 'directeur',
        'is_employee_view': role_slug not in ('admin', 'directeur_general', 'directeur') and not user.is_superuser,
        'total_budget': total_budget,
        'total_consumed': total_consumed,
        'budget_percentage': budget_percentage,
        'pending_documents': docs_qs.count(),
        'pending_requests': reqs_qs.count(),
        'active_partners': active_partners,
        'active_projects': active_projects,
        'pending_docs': pending_docs,
        'pending_reqs': pending_reqs,
        'upcoming_events': upcoming_events,
        'budget_data': budget_data,
        **project_stats,
        **role_context,
        **rbac_context,
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def projects(request):
    """Vue de la liste des projets avec restrictions par role."""
    status_filter = request.GET.get('status', 'all')
    direction_filter = request.GET.get('direction', 'all')
    search = request.GET.get('search', '')

    base_qs = get_accessible_projects_qs(request.user)
    projects_qs = base_qs.prefetch_related('milestones')

    today = _date.today()

    if status_filter == 'en_retard':
        projects_qs = projects_qs.filter(status__in=['planifie', 'en_cours'], end_date__lt=today)
    elif status_filter != 'all':
        projects_qs = projects_qs.filter(status=status_filter)
    if direction_filter != 'all':
        projects_qs = projects_qs.filter(direction__code=direction_filter)
    if search:
        projects_qs = projects_qs.filter(name__icontains=search)

    directions = Direction.objects.filter(projects__in=base_qs).distinct().order_by('name')

    stats = {
        'total': base_qs.count(),
        'en_cours': base_qs.filter(status='en_cours').count(),
        'termine': base_qs.filter(status='termine').count(),
        'planifie': base_qs.filter(status='planifie').count(),
        'suspendu': base_qs.filter(status='suspendu').count(),
        'en_retard': base_qs.filter(status__in=['planifie', 'en_cours'], end_date__lt=today).count(),
    }

    context = {
        'projects': projects_qs,
        'directions': directions,
        'stats': stats,
        'current_status': status_filter,
        'current_direction': direction_filter,
        'search': search,
    }
    return render(request, 'core/projects.html', context)


@login_required
def resources(request):
    """Vue des ressources (budget et RH)"""
    tab = request.GET.get('tab', 'budget')
    
    # Budget data
    can_view_budgets = request.user.profile.can_view_budgets()
    can_manage_budgets = request.user.profile.can_manage_budgets()
    can_view_all_budget_directions = request.user.profile.can_view_all_budget_directions()

    budgets_qs = Budget.objects.select_related('direction', 'project')
    if not can_view_budgets:
        budgets_qs = budgets_qs.none()
    elif not can_view_all_budget_directions:
        budgets_qs = budgets_qs.filter(Q(project__isnull=False) | Q(direction=request.user.profile.direction))

    from .currencies import convert_currency, format_currency
    from types import SimpleNamespace

    # 1) Budgets reels (table Budget)
    budgets = list(budgets_qs.all())

    # Projects deja couverts par une ligne Budget(project=...)
    projects_with_budget_row = {b.project_id for b in budgets if b.project_id}

    # 2) Synthese : Project.budget saisi directement sur le projet (sans ligne Budget dediee)
    if can_view_budgets:
        _inline_qs = get_accessible_projects_qs(request.user).exclude(id__in=projects_with_budget_row).filter(Q(budget__gt=0) | Q(budget_consumed__gt=0))
        for p in _inline_qs:
            budgets.append(SimpleNamespace(
                id=None,
                is_inline=True,  # source = Project.budget
                project=p,
                direction=p.direction,
                allocated=p.budget or 0,
                consumed=p.budget_consumed or 0,
                currency=p.currency or 'GNF',
                available=(p.budget or 0) - (p.budget_consumed or 0),
                consumption_rate=round((float(p.budget_consumed) / float(p.budget)) * 100, 1) if p.budget and p.budget > 0 else 0,
            ))

    # Conversion GNF + totaux
    total_allocated = 0
    total_consumed = 0
    for b in budgets:
        b.allocated_gnf = round(convert_currency(float(b.allocated), b.currency, 'GNF'))
        b.consumed_gnf = round(convert_currency(float(b.consumed), b.currency, 'GNF'))
        b.available_gnf = b.allocated_gnf - b.consumed_gnf
        total_allocated += b.allocated_gnf
        total_consumed += b.consumed_gnf
    total_allocated = round(total_allocated)
    total_consumed = round(total_consumed)
    
    # Employees data : scoped to direction for non-global users
    employees = Employee.objects.select_related('direction').all()
    profile = request.user.profile
    _emp_is_global = request.user.is_superuser or profile.is_admin() or profile.is_directeur_general()
    if not _emp_is_global and not profile.can_view_all_budget_directions():
        if profile.direction_id:
            employees = employees.filter(direction_id=profile.direction_id)
        else:
            employees = employees.none()

    # Stats globales calculées avant les filtres de recherche
    avg_workload = employees.aggregate(avg=Avg('workload'))['avg'] or 0
    overloaded = employees.filter(workload__gte=85).count()
    employees_count = employees.count()

    # Filtres de recherche RH (n'affectent que la liste, pas les stats)
    rh_search  = request.GET.get('search', '').strip()
    rh_dir     = request.GET.get('dir', '').strip()
    rh_workload = request.GET.get('workload', '').strip()
    if rh_search:
        employees = employees.filter(name__icontains=rh_search)
    if rh_dir:
        employees = employees.filter(direction_id=rh_dir)
    if rh_workload == 'surcharge':
        employees = employees.filter(workload__gte=85)

    directions = Direction.objects.all()
    
    # Employees linked to a user account (for badge display in template)
    employees_with_profile = set(
        UserProfile.objects.filter(employee__isnull=False).values_list('employee_id', flat=True)
    )

    context = {
        'tab': tab,
        'budgets': budgets,
        'total_allocated': total_allocated,
        'total_consumed': total_consumed,
        'total_available': total_allocated - total_consumed,
        'consumption_rate': round((float(total_consumed) / float(total_allocated) * 100), 1) if total_allocated > 0 else 0,
        'can_view_budgets': can_view_budgets,
        'can_manage_budgets': can_manage_budgets,
        'can_view_all_budget_directions': can_view_all_budget_directions,
        'employees': employees,
        'employees_count': employees_count,
        'avg_workload': round(avg_workload),
        'overloaded': overloaded,
        'directions': directions,
        'employees_with_profile': employees_with_profile,
        'rh_search': rh_search,
        'rh_dir': rh_dir,
        'rh_workload': rh_workload,
    }
    return render(request, 'core/resources.html', context)


@login_required
def documents(request):
    """Vue des documents"""
    profile = request.user.profile
    if not profile.can_approve_documents() and not profile.direction_id:
        messages.error(request, _("Vous n'avez pas accès aux documents."))
        return redirect('core:dashboard')

    status_filter = request.GET.get('status', 'all')
    type_filter = request.GET.get('type', 'all')
    search = request.GET.get('search', '')

    direction_id = getattr(profile, 'direction_id', None)

    if profile.is_directeur_general():
        docs_qs = Document.objects.select_related('direction')
    elif direction_id:
        docs_qs = Document.objects.select_related('direction').filter(direction_id=direction_id)
    else:
        docs_qs = Document.objects.none()

    stats = docs_qs.aggregate(
        total=Count('id'),
        a_signer=Count(Case(When(status='a_signer', then=1), output_field=IntegerField())),
        a_valider=Count(Case(When(status='a_valider', then=1), output_field=IntegerField())),
        signe=Count(Case(When(status='signe', then=1), output_field=IntegerField())),
    )

    if status_filter != 'all':
        docs_qs = docs_qs.filter(status=status_filter)
    if type_filter != 'all':
        docs_qs = docs_qs.filter(doc_type=type_filter)
    if search:
        docs_qs = docs_qs.filter(title__icontains=search)
    
    from django.utils import timezone
    context = {
        'documents': docs_qs,
        'stats': stats,
        'current_status': status_filter,
        'current_type': type_filter,
        'search': search,
        'today': timezone.now().date(),
    }
    return render(request, 'core/documents.html', context)


@login_required
def requests_view(request):
    """Vue des demandes"""
    status_filter = request.GET.get('status', 'all')
    direction_filter = request.GET.get('direction', 'all')
    search = request.GET.get('search', '')

    profile = request.user.profile
    direction_id = getattr(profile, 'direction_id', None)

    # DG ou rôle avec can_approve_requests → accès global
    if profile.is_directeur_general() or profile.can_approve_requests():
        reqs_qs = Request.objects.select_related('direction')
    elif direction_id:
        reqs_qs = Request.objects.select_related('direction').filter(direction_id=direction_id)
    else:
        reqs_qs = Request.objects.none()

    stats = reqs_qs.aggregate(
        total=Count('id'),
        en_attente=Count(Case(When(status='en_attente', then=1), output_field=IntegerField())),
        approuve=Count(Case(When(status='approuve', then=1), output_field=IntegerField())),
        rejete=Count(Case(When(status='rejete', then=1), output_field=IntegerField())),
    )

    if status_filter != 'all':
        reqs_qs = reqs_qs.filter(status=status_filter)
    if direction_filter != 'all':
        reqs_qs = reqs_qs.filter(direction__code=direction_filter)
    if search:
        reqs_qs = reqs_qs.filter(title__icontains=search)

    directions = Direction.objects.all()
    
    context = {
        'requests': reqs_qs,
        'directions': directions,
        'stats': stats,
        'current_status': status_filter,
        'current_direction': direction_filter,
        'search': search,
    }
    return render(request, 'core/requests.html', context)


@login_required
def calendar(request):
    """Vue du calendrier"""
    import calendar as cal

    if not request.user.profile.can_read_events():
        messages.error(request, _("Vous n'avez pas accès au calendrier."))
        return redirect('core:dashboard')

    today = timezone.now().date()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year, month = today.year, today.month
    if month < 1 or month > 12:
        month = today.month

    _cal_profile = request.user.profile
    _cal_is_global = _cal_profile.is_admin() or _cal_profile.is_directeur_general()
    _cal_dir_id = getattr(_cal_profile, 'direction_id', None)

    _base_events = Event.objects.prefetch_related('participants').select_related('created_by')
    if not _cal_is_global and _cal_dir_id:
        _base_events = _base_events.filter(participants__id=_cal_dir_id)
    elif not _cal_is_global:
        _base_events = _base_events.none()

    events_qs = _base_events.filter(date__year=year, date__month=month)

    stats = {
        'reunions': events_qs.filter(event_type='reunion').count(),
        'evenements': events_qs.filter(event_type='evenement').count(),
        'total': events_qs.count(),
    }

    upcoming = _base_events.filter(date__gte=today)[:5]

    cal_obj = cal.Calendar(firstweekday=0)
    month_days = cal_obj.monthdayscalendar(year, month)

    events_by_day = {}
    for event in events_qs:
        day = event.date.day
        if day not in events_by_day:
            events_by_day[day] = []
        events_by_day[day].append(event)

    months = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
              'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

    can_manage = request.user.profile.can_manage_events()

    context = {
        'year': year,
        'month': month,
        'month_name': months[month - 1],
        'month_days': month_days,
        'events_by_day': events_by_day,
        'upcoming_events': upcoming,
        'today': today,
        'prev_month': month - 1 if month > 1 else 12,
        'prev_year': year if month > 1 else year - 1,
        'next_month': month + 1 if month < 12 else 1,
        'next_year': year if month < 12 else year + 1,
        'stats': stats,
        'can_manage': can_manage,
    }
    return render(request, 'core/calendar.html', context)


@login_required
def event_detail(request, event_id):
    """Détail d'un événement"""
    event = get_object_or_404(Event, pk=event_id)

    if not request.user.profile.can_read_events():
        messages.error(request, _("Accès au calendrier insuffisant."))
        return redirect('core:dashboard')

    can_manage = request.user.profile.can_manage_events()
    is_creator = event.created_by == request.user
    can_edit = can_manage or is_creator

    members = (
        event.event_members
        .select_related('employee', 'employee__direction')
        .order_by('employee__name')
    )

    # Membership du user courant (pour RSVP)
    my_membership = None
    employee = getattr(request.user.profile, 'employee', None)
    if employee:
        my_membership = members.filter(employee=employee).first()

    # Compteurs RSVP
    rsvp_counts = {
        'accepte':   members.filter(status='accepte').count(),
        'refuse':    members.filter(status='refuse').count(),
        'peut_etre': members.filter(status='peut_etre').count(),
        'invite':    members.filter(status='invite').count(),
    }

    # Données pour l'ajout de membres (managers uniquement)
    existing_emp_ids = set(members.values_list('employee_id', flat=True))
    available_employees = Employee.objects.exclude(id__in=existing_emp_ids).order_by('name')
    all_directions = Direction.objects.order_by('name')
    all_projects = get_accessible_projects_qs(request.user).order_by('name')

    return render(request, 'core/event_detail.html', {
        'event': event,
        'can_edit': can_edit,
        'can_delete': can_edit,
        'members': members,
        'my_membership': my_membership,
        'rsvp_counts': rsvp_counts,
        'available_employees': available_employees,
        'all_directions': all_directions,
        'all_projects': all_projects,
    })


def _sync_event_members(event, actor_user):
    """
    Synchronise les EventMember selon les directions participantes (dans une transaction).
    - Bulk-crée les membres manquants et les notifie (invite)
    - Retire les membres non-répondants dont la direction est retirée
    Retourne le set d'employee_id nouvellement créés.
    """
    from django.db import transaction
    from .notifs import notify_event_member_invited
    from .notifications import notify_event_member_email

    with transaction.atomic():
        current_dir_ids = set(event.participants.values_list('id', flat=True))
        existing = {
            m.employee_id: m
            for m in event.event_members.select_related('employee').all()
        }
        expected_ids = set(
            Employee.objects
            .filter(direction_id__in=current_dir_ids)
            .values_list('id', flat=True)
        )
        to_add_ids = expected_ids - set(existing.keys())

        if to_add_ids:
            employees_to_add = Employee.objects.filter(id__in=to_add_ids)
            EventMember.objects.bulk_create(
                [EventMember(event=event, employee=emp) for emp in employees_to_add],
                ignore_conflicts=True,
            )
            # Récupérer les instances créées (bulk_create ne remplit pas les PK sur SQLite)
            for member in EventMember.objects.filter(event=event, employee_id__in=to_add_ids).select_related('employee'):
                notify_event_member_invited(member, actor_user)
                notify_event_member_email(member, actor_user)

        # Retirer les membres non-répondants dont la direction est retirée
        to_remove = [
            m for emp_id, m in existing.items()
            if emp_id not in expected_ids and m.status == 'invite'
        ]
        for m in to_remove:
            m.delete()

    return to_add_ids


@login_required
def event_create(request):
    """Créer un événement"""
    from .forms import EventForm
    from .notifs import push_calendar_update

    if not request.user.profile.can_manage_events():
        messages.error(request, _("Permission insuffisante pour créer des événements."))
        return redirect('core:calendar')

    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            event.created_by = request.user
            event.save()
            form.save_m2m()
            # _sync_event_members envoie les invitations individuelles — pas de notif direction séparée
            _sync_event_members(event, request.user)
            push_calendar_update('created', event.id, event.date)
            messages.success(request, _("Événement créé avec succès."))
            return redirect('core:event_detail', event_id=event.pk)
    else:
        initial = {}
        date_str = request.GET.get('date')
        if date_str:
            initial['date'] = date_str
        form = EventForm(initial=initial)

    return render(request, 'core/event_form.html', {'form': form, 'title': 'Nouvel événement'})


@login_required
def event_edit(request, event_id):
    """Modifier un événement"""
    from .forms import EventForm
    from .notifs import notify_event_members_updated, push_calendar_update

    event = get_object_or_404(Event, pk=event_id)
    can_manage = request.user.profile.can_manage_events()
    is_creator = event.created_by == request.user

    if not can_manage and not is_creator:
        messages.error(request, _("Vous ne pouvez modifier que vos propres événements."))
        return redirect('core:event_detail', event_id=event_id)

    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            form.save_m2m()
            # new_ids = membres nouvellement invités via sync (ont déjà reçu une notif invite)
            new_ids = _sync_event_members(event, request.user)
            # Notifier uniquement les membres existants (pas les nouveaux) de la modification
            notify_event_members_updated(event, request.user, exclude_employee_ids=new_ids)
            push_calendar_update('updated', event.id, event.date)
            messages.success(request, _("Événement modifié avec succès."))
            return redirect('core:event_detail', event_id=event.pk)
    else:
        form = EventForm(instance=event)

    return render(request, 'core/event_form.html', {
        'form': form, 'event': event, 'title': f'Modifier — {event.title}'
    })


@login_required
def event_delete(request, event_id):
    """Supprimer un événement"""
    from .notifs import notify_event_deleted, push_calendar_update

    event = get_object_or_404(Event, pk=event_id)
    can_manage = request.user.profile.can_manage_events()
    is_creator = event.created_by == request.user

    if not can_manage and not is_creator:
        messages.error(request, _("Vous ne pouvez supprimer que vos propres événements."))
        return redirect('core:event_detail', event_id=event_id)

    if request.method == 'POST':
        from .notifications import notify_event_deleted_email
        directions = list(event.participants.all())
        event_title, event_date = event.title, event.date
        # Capturer les emails AVANT suppression (CASCADE supprime les EventMember)
        member_emails = list(
            event.event_members
            .exclude(employee__email='')
            .values_list('employee__email', flat=True)
        )
        event.delete()
        notify_event_deleted(event_title, event_date, directions, request.user)
        notify_event_deleted_email(event_title, event_date, member_emails, request.user)
        push_calendar_update('deleted', None, event_date)
        messages.success(request, _("Événement supprimé."))
        return redirect('core:calendar')

    return render(request, 'core/event_confirm_delete.html', {'event': event})


@login_required
def event_rsvp(request, event_id):
    """L'utilisateur répond à son invitation (RSVP)."""
    from .models import EventMember
    from .notifs import notify_event_rsvp

    if request.method != 'POST':
        return redirect('core:event_detail', event_id=event_id)

    event = get_object_or_404(Event, pk=event_id)
    employee = getattr(request.user.profile, 'employee', None)
    if not employee:
        messages.error(request, _("Votre profil n'est pas lié à un employé."))
        return redirect('core:event_detail', event_id=event_id)

    member = get_object_or_404(EventMember, event=event, employee=employee)
    status = request.POST.get('status')
    if status not in ('accepte', 'refuse', 'peut_etre'):
        messages.error(request, _("Statut invalide."))
        return redirect('core:event_detail', event_id=event_id)

    member.status = status
    member.note = request.POST.get('note', '')
    member.responded_at = timezone.now()
    member.save(update_fields=['status', 'note', 'responded_at'])
    notify_event_rsvp(member, request.user)
    messages.success(request, _("Votre réponse (%(status)s) a été enregistrée.") % {
        'status': member.get_status_display()
    })
    return redirect('core:event_detail', event_id=event_id)


def _event_can_manage_members(request, event):
    """True si l'utilisateur peut gérer les membres de cet événement."""
    profile = request.user.profile
    return profile.can_manage_events() or event.created_by == request.user


@login_required
def event_add_members(request, event_id):
    """Ajouter des membres à un événement : individuel, direction entière, projet entier."""
    from .notifs import notify_event_member_invited
    from .notifications import notify_event_member_email

    if request.method != 'POST':
        return redirect('core:event_detail', event_id=event_id)

    event = get_object_or_404(Event, pk=event_id)
    if not _event_can_manage_members(request, event):
        messages.error(request, _("Permission insuffisante."))
        return redirect('core:event_detail', event_id=event_id)

    add_type = request.POST.get('add_type', '')
    added = 0

    def _add_employee(emp):
        nonlocal added
        member, created = EventMember.objects.get_or_create(event=event, employee=emp)
        if created:
            notify_event_member_invited(member, request.user)
            notify_event_member_email(member, request.user)
            added += 1

    if add_type == 'employee':
        emp_id = request.POST.get('employee_id')
        if emp_id:
            emp = get_object_or_404(Employee, pk=emp_id)
            _add_employee(emp)

    elif add_type == 'direction':
        dir_id = request.POST.get('direction_id')
        if dir_id:
            for emp in Employee.objects.filter(direction_id=dir_id):
                _add_employee(emp)

    elif add_type == 'project':
        proj_id = request.POST.get('project_id')
        if proj_id:
            for pm in ProjectMember.objects.filter(project_id=proj_id).select_related('employee'):
                _add_employee(pm.employee)

    if added:
        messages.success(request, _("%(n)d participant(s) ajouté(s).") % {'n': added})
    else:
        messages.info(request, _("Tous ces participants sont déjà invités."))
    return redirect('core:event_detail', event_id=event_id)


@login_required
def event_remove_member(request, event_id, member_id):
    """Retirer un membre d'un événement."""
    if request.method != 'POST':
        return redirect('core:event_detail', event_id=event_id)

    event = get_object_or_404(Event, pk=event_id)
    if not _event_can_manage_members(request, event):
        messages.error(request, _("Permission insuffisante."))
        return redirect('core:event_detail', event_id=event_id)

    member = get_object_or_404(EventMember, pk=member_id, event=event)
    emp_name = member.employee.name
    member.delete()
    messages.success(request, _("%(name)s retiré(e) de l'événement.") % {'name': emp_name})
    return redirect('core:event_detail', event_id=event_id)


@login_required
def reports(request):
    """Vue des rapports"""
    if not request.user.profile.can_view_reports():
        messages.error(request, _("Accès insuffisant pour consulter les rapports."))
        return redirect('core:dashboard')
    
    # Périmètre selon le rôle
    _scoped_projects = get_accessible_projects_qs(request.user)

    # Projets par direction (uniquement celles accessibles)
    directions = Direction.objects.filter(projects__in=_scoped_projects).distinct().annotate(
        proj_en_cours=Count(Case(When(projects__status='en_cours', projects__in=_scoped_projects, then=1), output_field=IntegerField())),
        proj_termine=Count(Case(When(projects__status='termine', projects__in=_scoped_projects, then=1), output_field=IntegerField())),
    )
    projects_by_direction = [
        {'direction': d.code, 'en_cours': d.proj_en_cours, 'termine': d.proj_termine}
        for d in directions
    ]

    # Projets par statut + KPIs — scopés
    proj_agg = _scoped_projects.aggregate(
        total=Count('id'),
        en_cours=Count(Case(When(status='en_cours', then=1), output_field=IntegerField())),
        termine=Count(Case(When(status='termine', then=1), output_field=IntegerField())),
        planifie=Count(Case(When(status='planifie', then=1), output_field=IntegerField())),
    )
    projects_by_status = [
        {'name': 'En cours', 'value': proj_agg['en_cours']},
        {'name': 'Terminés', 'value': proj_agg['termine']},
        {'name': 'Planifiés', 'value': proj_agg['planifie']},
    ]
    total_projects = proj_agg['total']
    projects_completed = proj_agg['termine']
    _accessible_proj_ids = set(_scoped_projects.values_list('id', flat=True))
    budget_agg = Budget.objects.filter(
        Q(project_id__in=_accessible_proj_ids) | Q(direction__in=directions)
    ).aggregate(alloc=Sum('allocated'), cons=Sum('consumed'))
    _budget_table_alloc = float(budget_agg['alloc'] or 0)
    _budget_table_cons  = float(budget_agg['cons']  or 0)
    _project_ids_with_budget_row = set(
        Budget.objects.filter(project_id__in=_accessible_proj_ids).values_list('project_id', flat=True)
    )
    _project_direct = _scoped_projects.exclude(id__in=_project_ids_with_budget_row).aggregate(
        a=Sum('budget'), c=Sum('budget_consumed')
    )
    total_budget   = _budget_table_alloc + float(_project_direct['a'] or 0)
    total_consumed = _budget_table_cons  + float(_project_direct['c'] or 0)
    budget_rate = round((total_consumed / total_budget * 100)) if total_budget > 0 else 0
    active_partners = Partner.objects.filter(status='actif').count()
    _profile = request.user.profile
    if _profile.direction_id and not (_profile.is_admin() or _profile.is_directeur_general()):
        pending_docs = Document.objects.filter(direction_id=_profile.direction_id).exclude(status='signe').count()
    else:
        pending_docs = Document.objects.exclude(status='signe').count()

    # Tous les projets accessibles pour le tableau
    all_projects = _scoped_projects.select_related('direction')

    # Budget par direction : Budget(direction) + Budget(project__direction) + Project.budget direct
    budget_by_direction = []
    for direction in directions:
        # Source 1 : Budget rows linked directly to the direction
        dir_budget_agg = direction.budgets.aggregate(a=Sum('allocated'), c=Sum('consumed'))
        dir_alloc = float(dir_budget_agg['a'] or 0)
        dir_cons  = float(dir_budget_agg['c'] or 0)

        # Source 2 : Budget rows linked to accessible projects in this direction
        proj_budget_agg = Budget.objects.filter(
            project_id__in=_accessible_proj_ids, project__direction=direction
        ).aggregate(a=Sum('allocated'), c=Sum('consumed'))
        dir_alloc += float(proj_budget_agg['a'] or 0)
        dir_cons  += float(proj_budget_agg['c'] or 0)

        # Source 3 : Project.budget for accessible projects not covered by a Budget row
        covered_ids = set(
            Budget.objects.filter(project_id__in=_accessible_proj_ids, project__direction=direction).values_list('project_id', flat=True)
        )
        proj_direct_agg = _scoped_projects.filter(direction=direction).exclude(
            id__in=covered_ids
        ).aggregate(a=Sum('budget'), c=Sum('budget_consumed'))
        dir_alloc += float(proj_direct_agg['a'] or 0)
        dir_cons  += float(proj_direct_agg['c'] or 0)

        budget_by_direction.append({
            'direction': direction.code,
            'name': direction.name,
            'allocated': round(dir_alloc),
            'consumed': round(dir_cons),
            'rate': round((dir_cons / dir_alloc * 100)) if dir_alloc > 0 else 0,
        })
    
    # Performance par projet (scopé)
    projects_en_retard = _scoped_projects.filter(
        status__in=['planifie', 'en_cours'], end_date__lt=_date.today()
    ).count()
    projects_en_cours = _scoped_projects.filter(status='en_cours').count()
    avg_progress = all_projects.aggregate(avg=Avg('progress'))['avg'] or 0
    
    # Demandes en attente (scopées à la direction pour les non-DG)
    if _profile.direction_id and not (_profile.is_admin() or _profile.is_directeur_general()):
        pending_requests = Request.objects.filter(status='en_attente', direction_id=_profile.direction_id).count()
    else:
        pending_requests = Request.objects.filter(status='en_attente').count()

    context = {
        'projects_by_direction': projects_by_direction,
        'projects_by_status': projects_by_status,
        'total_projects': total_projects,
        'projects_completed': projects_completed,
        'budget_rate': budget_rate,
        'active_partners': active_partners,
        'pending_docs': pending_docs,
        'pending_requests': pending_requests,
        'all_projects': all_projects,
        'total_budget': total_budget,
        'total_consumed': total_consumed,
        'budget_by_direction': budget_by_direction,
        'projects_en_retard': projects_en_retard,
        'projects_en_cours': projects_en_cours,
        'avg_progress': round(avg_progress),
    }
    return render(request, 'core/reports.html', context)


@login_required
def export_employees_pdf(request):
    """Exporter la liste des employés en PDF"""
    _ep = request.user.profile
    if not _ep.is_directeur():
        messages.error(request, _("Accès réservé aux directeurs et administrateurs."))
        return redirect('core:resources')
    employees = Employee.objects.select_related('direction').all()
    if _ep.direction_id and not (_ep.is_admin() or _ep.is_directeur_general()):
        employees = employees.filter(direction_id=_ep.direction_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="employes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []
    
    title_style = ParagraphStyle(
        'EmpTitle',
        parent=styles['Heading1'],
        fontSize=22,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'EmpSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.grey
    )
    
    heading_style = ParagraphStyle(
        'EmpHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10,
        spaceBefore=15,
        textColor=colors.darkblue
    )
    
    # En-tête
    story.append(Paragraph("LISTE DES EMPLOYÉS", title_style))
    story.append(Paragraph(f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    story.append(Paragraph(f"Effectif total : {employees.count()} employés", subtitle_style))
    story.append(Spacer(1, 20))
    
    # Tableau des employés
    story.append(Paragraph("RÉPERTOIRE DU PERSONNEL", heading_style))
    
    table_data = [['Nom', 'Direction', 'Rôle', 'Téléphone', 'Email', 'Charge']]
    
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10)
    header_cell_style = ParagraphStyle('HeaderCell', parent=styles['Normal'], fontSize=9, leading=11, textColor=colors.whitesmoke)
    
    table_data = [[
        Paragraph('<b>Nom</b>', header_cell_style),
        Paragraph('<b>Direction</b>', header_cell_style),
        Paragraph('<b>Rôle</b>', header_cell_style),
        Paragraph('<b>Téléphone</b>', header_cell_style),
        Paragraph('<b>Email</b>', header_cell_style),
        Paragraph('<b>Charge</b>', header_cell_style),
    ]]
    
    for emp in employees:
        table_data.append([
            Paragraph(emp.name, cell_style),
            Paragraph(emp.direction.code if emp.direction else '-', cell_style),
            Paragraph(emp.role, cell_style),
            Paragraph(emp.phone or '-', cell_style),
            Paragraph(emp.email or '-', cell_style),
            Paragraph(f'{emp.workload}%', cell_style),
        ])
    
    col_widths = [1.4*inch, 0.8*inch, 1.3*inch, 1.0*inch, 1.6*inch, 0.6*inch]
    emp_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (-1, 1), (-1, -1), 'CENTER'),
    ]
    
    # Alterner les couleurs de fond
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.Color(0.95, 0.95, 0.97)))
    
    emp_table.setStyle(TableStyle(table_style))
    story.append(emp_table)
    story.append(Spacer(1, 20))
    
    # Résumé par direction
    story.append(Paragraph("RÉPARTITION PAR DIRECTION", heading_style))
    
    dir_data = [[
        Paragraph('<b>Direction</b>', header_cell_style),
        Paragraph('<b>Effectif</b>', header_cell_style),
        Paragraph('<b>Charge moyenne</b>', header_cell_style),
    ]]
    
    directions = Direction.objects.filter(employees__in=employees).distinct()
    for d in directions:
        dir_employees = employees.filter(direction=d)
        count = dir_employees.count()
        if count > 0:
            avg_wl = round(sum(e.workload for e in dir_employees) / count)
            dir_data.append([
                Paragraph(f'{d.code} - {d.name}', cell_style),
                Paragraph(str(count), cell_style),
                Paragraph(f'{avg_wl}%', cell_style),
            ])
    
    dir_table = Table(dir_data, colWidths=[3*inch, 1.2*inch, 1.5*inch], repeatRows=1)
    dir_table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    for i in range(1, len(dir_data)):
        if i % 2 == 0:
            dir_table_style.append(('BACKGROUND', (0, i), (-1, i), colors.Color(0.95, 0.95, 0.97)))
    
    dir_table.setStyle(TableStyle(dir_table_style))
    story.append(dir_table)
    
    doc.build(story)
    return response


@login_required
def export_reports_pdf(request):
    """Exporter les rapports en PDF"""
    if not request.user.profile.can('export', 'Report'):
        messages.error(request, _("Vous n'avez pas la permission d'exporter les rapports."))
        return redirect('core:reports')

    # Périmètre selon le rôle (même logique que la vue reports)
    _scoped = get_accessible_projects_qs(request.user)
    directions = Direction.objects.filter(projects__in=_scoped).distinct().annotate(
        proj_en_cours=Count(Case(When(projects__status='en_cours', projects__in=_scoped, then=1), output_field=IntegerField())),
        proj_termine=Count(Case(When(projects__status='termine', projects__in=_scoped, then=1), output_field=IntegerField())),
    )
    projects_by_direction = [
        {'direction': d.code, 'en_cours': d.proj_en_cours, 'termine': d.proj_termine}
        for d in directions
    ]

    proj_agg = _scoped.aggregate(
        total=Count('id'),
        en_cours=Count(Case(When(status='en_cours', then=1), output_field=IntegerField())),
        termine=Count(Case(When(status='termine', then=1), output_field=IntegerField())),
        planifie=Count(Case(When(status='planifie', then=1), output_field=IntegerField())),
    )
    projects_by_status = [
        {'name': 'En cours', 'value': proj_agg['en_cours']},
        {'name': 'Terminés', 'value': proj_agg['termine']},
        {'name': 'Planifiés', 'value': proj_agg['planifie']},
    ]
    total_projects = proj_agg['total']
    projects_completed = proj_agg['termine']
    _scoped_ids = set(_scoped.values_list('id', flat=True))
    budget_agg = Budget.objects.filter(
        Q(project_id__in=_scoped_ids) | Q(direction__in=directions)
    ).aggregate(total=Sum('allocated'), consumed=Sum('consumed'))
    total_budget = float(budget_agg['total'] or 0)
    total_consumed = float(budget_agg['consumed'] or 0)
    budget_rate = round((total_consumed / total_budget * 100)) if total_budget > 0 else 0
    active_partners = Partner.objects.filter(status='actif').count()
    _exp_profile = request.user.profile
    if _exp_profile.direction_id and not (_exp_profile.is_admin() or _exp_profile.is_directeur_general()):
        pending_docs = Document.objects.filter(direction_id=_exp_profile.direction_id).exclude(status='signe').count()
    else:
        pending_docs = Document.objects.exclude(status='signe').count()
    all_projects = _scoped.select_related('direction')
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="rapports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    styles = getSampleStyleSheet()
    story = []
    
    # Style personnalisé pour le titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    # Style pour les sous-titres
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.darkblue
    )
    
    # En-tête
    story.append(Paragraph("RAPPORT D'ACTIVITÉS", title_style))
    story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Section KPIs
    story.append(Paragraph("INDICATEURS CLÉS", heading_style))
    
    kpi_data = [
        ['Indicateur', 'Valeur'],
        ['Projets totaux', str(total_projects)],
        ['Projets terminés', str(projects_completed)],
        ['Budget consommé', f'{budget_rate}%'],
        ['Partenaires actifs', str(active_partners)],
        ['Documents en attente', str(pending_docs)],
    ]
    
    kpi_table = Table(kpi_data, colWidths=[3*inch, 1.5*inch])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # Section Projets par direction
    story.append(Paragraph("PROJETS PAR DIRECTION", heading_style))
    
    direction_data = [['Direction', 'En cours', 'Terminés']]
    for item in projects_by_direction:
        direction_data.append([item['direction'], str(item['en_cours']), str(item['termine'])])
    
    direction_table = Table(direction_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
    direction_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(direction_table)
    story.append(Spacer(1, 20))
    
    # Section Projets par statut
    story.append(Paragraph("RÉPARTITION DES PROJETS PAR STATUT", heading_style))
    
    status_data = [['Statut', 'Nombre']]
    for item in projects_by_status:
        status_data.append([item['name'], str(item['value'])])
    
    status_table = Table(status_data, colWidths=[2.5*inch, 1.5*inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(status_table)
    story.append(Spacer(1, 20))
    
    # Section détaillée des projets
    story.append(PageBreak())
    story.append(Paragraph("DÉTAIL DES PROJETS", heading_style))
    
    projects_data = [['Projet', 'Direction', 'Statut', 'Progression', 'Budget', 'Consommé', 'Taux']]
    for project in all_projects:
        projects_data.append([
            project.name,
            project.direction.code if project.direction else '-',
            project.get_status_display(),
            f"{project.progress}%",
            f"{project.budget:,.0f} GNF",
            f"{project.budget_consumed:,.0f} GNF",
            f"{project.budget_percentage}%"
        ])
    
    projects_table = Table(projects_data, colWidths=[2*inch, 1*inch, 1*inch, 0.8*inch, 1*inch, 1*inch, 0.8*inch])
    projects_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Aligner les noms de projets à gauche
    ]))
    story.append(projects_table)
    
    # Construire le PDF
    doc.build(story)
    
    return response


@login_required
def partners(request):
    """Vue des partenaires"""
    if not request.user.profile.can_read_partners():
        messages.error(request, _("Vous n'avez pas la permission de consulter les partenaires."))
        return redirect('core:dashboard')
    type_filter = request.GET.get('type', 'all')
    status_filter = request.GET.get('status', 'all')
    search = request.GET.get('search', '')
    
    partners_qs = Partner.objects.all()
    
    if type_filter != 'all':
        partners_qs = partners_qs.filter(partner_type=type_filter)
    if status_filter != 'all':
        partners_qs = partners_qs.filter(status=status_filter)
    if search:
        partners_qs = partners_qs.filter(name__icontains=search)
    
    today = timezone.now().date()
    upcoming_events = Event.objects.filter(
        event_type='evenement',
        date__gte=today
    ).prefetch_related('participants')[:3]
    
    partner_agg = Partner.objects.aggregate(
        total=Count('id'),
        actif=Count(Case(When(status='actif', then=1), output_field=IntegerField())),
        en_discussion=Count(Case(When(status='en_discussion', then=1), output_field=IntegerField())),
        entreprise=Count(Case(When(partner_type='entreprise', then=1), output_field=IntegerField())),
        universite=Count(Case(When(partner_type='universite', then=1), output_field=IntegerField())),
        institution=Count(Case(When(partner_type='institution', then=1), output_field=IntegerField())),
        ong=Count(Case(When(partner_type='ong', then=1), output_field=IntegerField())),
    )
    stats = {
        'total': partner_agg['total'],
        'actif': partner_agg['actif'],
        'en_discussion': partner_agg['en_discussion'],
    }
    type_counts = {
        'entreprise': partner_agg['entreprise'],
        'universite': partner_agg['universite'],
        'institution': partner_agg['institution'],
        'ong': partner_agg['ong'],
    }
    
    context = {
        'partners': partners_qs,
        'stats': stats,
        'type_counts': type_counts,
        'upcoming_events': upcoming_events,
        'current_type': type_filter,
        'current_status': status_filter,
        'search': search,
    }
    return render(request, 'core/partners.html', context)


# ==================== DIRECTIONS CRUD ====================

@login_required
def directions_list(request):
    """Liste des directions"""
    # Seuls admin et DG peuvent gérer les directions
    if not (request.user.profile.is_directeur_general() or request.user.profile.is_admin()):
        messages.error(request, _("Vous n'avez pas la permission de gérer les directions."))
        return redirect('core:dashboard')
    
    directions = Direction.objects.annotate(
        projects_count=Count('projects', distinct=True),
        users_count=Count('users', distinct=True)
    ).order_by('name')

    totals = directions.aggregate(
        total_projects=Sum('projects_count'),
        total_users=Sum('users_count'),
    )

    context = {
        'directions': directions,
        'total_projects': totals['total_projects'] or 0,
        'total_users': totals['total_users'] or 0,
    }
    return render(request, 'core/directions_list.html', context)


@login_required
def direction_create(request):
    """Créer une nouvelle direction"""
    if not (request.user.profile.is_directeur_general() or request.user.profile.is_admin()):
        messages.error(request, _("Vous n'avez pas la permission de créer des directions."))
        return redirect('core:dashboard')
    
    from .forms import DirectionForm
    
    if request.method == 'POST':
        form = DirectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _("Direction créée avec succès."))
            return redirect('core:directions_list')
    else:
        form = DirectionForm()
    
    context = {
        'form': form,
        'title': 'Nouvelle direction',
    }
    return render(request, 'core/direction_form.html', context)


@login_required
def direction_edit(request, direction_id):
    """Modifier une direction"""
    if not (request.user.profile.is_directeur_general() or request.user.profile.is_admin()):
        messages.error(request, _("Vous n'avez pas la permission de modifier les directions."))
        return redirect('core:dashboard')
    
    direction = get_object_or_404(Direction, pk=direction_id)
    from .forms import DirectionForm
    
    if request.method == 'POST':
        form = DirectionForm(request.POST, instance=direction)
        if form.is_valid():
            form.save()
            messages.success(request, _("Direction modifiée avec succès."))
            return redirect('core:directions_list')
    else:
        form = DirectionForm(instance=direction)
    
    context = {
        'form': form,
        'direction': direction,
        'title': f'Modifier {direction.name}',
    }
    return render(request, 'core/direction_form.html', context)


@login_required
def direction_delete(request, direction_id):
    """Supprimer une direction"""
    if not (request.user.profile.is_directeur_general() or request.user.profile.is_admin()):
        messages.error(request, _("Vous n'avez pas la permission de supprimer les directions."))
        return redirect('core:dashboard')
    
    direction = get_object_or_404(Direction, pk=direction_id)
    
    if request.method == 'POST':
        direction.delete()
        messages.success(request, _("Direction supprimée."))
        return redirect('core:directions_list')
    
    return render(request, 'core/confirm_delete.html', {
        'object': direction,
        'type': 'direction',
        'back_url': 'core:directions_list'
    })


# ==================== USER MANAGEMENT ====================

@login_required
def users_list(request):
    """Liste des utilisateurs"""
    from django.contrib.auth.models import User
    from .models import UserActivity
    
    # Check permission
    if not request.user.profile.has_manage_users_permission():
        messages.error(request, _("Vous n'avez pas les permissions pour gérer les utilisateurs."))
        return redirect('core:dashboard')
    
    search = request.GET.get('search', '')
    role_filter = request.GET.get('role', 'all')
    status_filter = request.GET.get('status', 'all')
    
    _ul_profile = request.user.profile
    _ul_is_global = request.user.is_superuser or _ul_profile.is_admin() or _ul_profile.is_directeur_general()

    users = User.objects.select_related('profile', 'profile__employee', 'profile__direction')
    if not _ul_is_global and _ul_profile.direction_id:
        users = users.filter(profile__direction_id=_ul_profile.direction_id)

    if search:
        users = users.filter(
            models.Q(username__icontains=search) |
            models.Q(first_name__icontains=search) |
            models.Q(last_name__icontains=search) |
            models.Q(email__icontains=search)
        )

    if role_filter != 'all':
        users = users.filter(profile__role__slug=role_filter)

    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)

    # Stats (scoped identically)
    _ul_base = User.objects if _ul_is_global else User.objects.filter(profile__direction_id=_ul_profile.direction_id)
    stats = {
        'total': _ul_base.count(),
        'active': _ul_base.filter(is_active=True).count(),
        'inactive': _ul_base.filter(is_active=False).count(),
    }
    
    # Recent activities
    recent_activities = UserActivity.objects.select_related('user').all()[:10]
    
    context = {
        'users': users,
        'stats': stats,
        'recent_activities': recent_activities,
        'search': search,
        'current_role': role_filter,
        'current_status': status_filter,
        'role_choices': [(r.slug, r.name) for r in Role.objects.all().order_by('name')],
    }
    return render(request, 'core/users/list.html', context)


@login_required
def user_create(request):
    """Crée un compte inactif et envoie un email d'invitation à l'employé."""
    from .forms import UserCreateForm
    from django.contrib.auth.models import User as DjangoUser
    import json

    if not request.user.profile.has_manage_users_permission():
        messages.error(request, _("Vous n'avez pas les permissions pour créer des utilisateurs."))
        return redirect('core:dashboard')

    initial = {}
    employee_id = request.GET.get('employee_id')
    if employee_id:
        try:
            emp = Employee.objects.get(pk=employee_id)
            initial['employee'] = emp
            initial['role'] = 'visiteur' if emp.is_external else 'employe'
        except Employee.DoesNotExist:
            pass

    if request.method == 'POST':
        form = UserCreateForm(request.POST, editor=request.user)
        if form.is_valid():
            emp = form.cleaned_data['employee']
            role = form.cleaned_data['role']

            username = _generate_username(emp.name)
            name_parts = emp.name.strip().split()

            user = DjangoUser(
                username=username,
                email=emp.email,
                first_name=name_parts[0] if name_parts else '',
                last_name=' '.join(name_parts[1:]) if len(name_parts) > 1 else '',
                is_active=False,
            )
            user.set_unusable_password()
            user.save()

            profile, _created = UserProfile.objects.get_or_create(user=user)
            profile.role = role  # role est déjà un objet Role (ModelChoiceField)
            profile.direction = emp.direction
            profile.employee = emp
            profile.save()

            activation_link = _build_activation_link(request, user)
            invited_by = request.user.get_full_name() or request.user.username
            from .notifications import notify_account_invitation
            email_ok, email_msg = notify_account_invitation(user, activation_link, invited_by)

            UserActivity.objects.create(
                user=request.user,
                action='create',
                description=f"Création du compte {username} pour {emp.name}",
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
            )

            if email_ok:
                messages.success(request, _("Compte créé pour {name}. Invitation envoyée à {email}.").format(name=emp.name, email=emp.email))
            else:
                messages.warning(request, _("Compte créé (identifiant : {username}) mais l'email n'a pas pu être envoyé : {msg}").format(username=username, msg=email_msg))
            return redirect('core:users_list')
    else:
        form = UserCreateForm(initial=initial, editor=request.user)

    # Données JSON pour la prévisualisation JS (username, direction, email)
    import unicodedata, re as _re

    def _norm(s):
        s = unicodedata.normalize('NFD', s)
        s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
        return _re.sub(r'[^a-z0-9]', '', s.lower())

    emp_preview = {}
    for emp in form.fields['employee'].queryset:
        parts = emp.name.strip().split()
        prenom = _norm(parts[0]) if parts else ''
        nom = _norm(parts[-1]) if len(parts) >= 2 else prenom
        emp_preview[str(emp.id)] = {
            'name': emp.name,
            'email': emp.email or '',
            'direction_code': emp.direction.code if emp.direction else '',
            'direction_name': emp.direction.name if emp.direction else '',
            'preview_username': f"{nom}.{prenom}" if prenom and nom != prenom else nom,
        }

    return render(request, 'core/users/create.html', {
        'form': form,
        'emp_preview_json': json.dumps(emp_preview),
        'title': 'Nouveau compte utilisateur',
    })


@login_required
def user_edit(request, user_id):
    """Modifier un utilisateur"""
    from django.contrib.auth.models import User
    from .forms import UserUpdateForm
    
    if not request.user.profile.has_manage_users_permission():
        messages.error(request, _("Vous n'avez pas les permissions pour modifier des utilisateurs."))
        return redirect('core:dashboard')
    
    user_obj = get_object_or_404(User, pk=user_id)
    
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=user_obj, editor=request.user)
        if form.is_valid():
            # Bloquer l'activation d'un compte en attente d'invitation
            if not user_obj.has_usable_password() and not user_obj.is_active:
                form.instance.is_active = False

            # Identité verrouillée sur la fiche employé si liée
            employee = form.cleaned_data.get('employee') or (
                user_obj.profile.employee if hasattr(user_obj, 'profile') else None
            )
            if employee:
                name_parts = employee.name.split() if employee.name else []
                form.instance.first_name = name_parts[0] if name_parts else ''
                form.instance.last_name  = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                form.instance.email      = employee.email or user_obj.email

            form.save()

            # Log activity
            UserActivity.objects.create(
                user=request.user,
                action='update',
                description=f"Modification de l'utilisateur {user_obj.username}",
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
            )

            messages.success(request, _("L'utilisateur {username} a été modifié avec succès.").format(username=user_obj.username))
            return redirect('core:users_list')
    else:
        form = UserUpdateForm(instance=user_obj, editor=request.user)
    
    context = {
        'form': form,
        'user_obj': user_obj,
        'title': f'Modifier {user_obj.username}',
    }
    return render(request, 'core/users/form.html', context)


@login_required
def user_delete(request, user_id):
    """Supprimer un utilisateur"""
    from django.contrib.auth.models import User
    
    if not request.user.profile.has_manage_users_permission():
        messages.error(request, _("Vous n'avez pas les permissions pour supprimer des utilisateurs."))
        return redirect('core:dashboard')
    
    user_obj = get_object_or_404(User, pk=user_id)

    if user_obj == request.user:
        messages.error(request, _("Vous ne pouvez pas supprimer votre propre compte."))
        return redirect('core:users_list')

    _ud_profile = request.user.profile
    _ud_is_global = request.user.is_superuser or _ud_profile.is_admin() or _ud_profile.is_directeur_general()
    if not _ud_is_global:
        _target_profile = getattr(user_obj, 'profile', None)
        if not _target_profile or _target_profile.direction_id != _ud_profile.direction_id:
            messages.error(request, _("Vous ne pouvez supprimer que les utilisateurs de votre direction."))
            return redirect('core:users_list')

    if request.method == 'POST':
        username = user_obj.username
        user_obj.delete()
        
        # Log activity
        UserActivity.objects.create(
            user=request.user,
            action='delete',
            description=f"Suppression de l'utilisateur {username}",
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
        )
        
        messages.success(request, _("L'utilisateur {username} a été supprimé.").format(username=username))
        return redirect('core:users_list')
    
    context = {
        'user_obj': user_obj,
    }
    return render(request, 'core/users/delete.html', context)


@login_required
def user_toggle_status(request, user_id):
    """Activer/Désactiver un utilisateur"""
    from django.contrib.auth.models import User
    
    if not request.user.profile.has_manage_users_permission():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:dashboard')
    
    user_obj = get_object_or_404(User, pk=user_id)

    if user_obj == request.user:
        messages.error(request, _("Vous ne pouvez pas désactiver votre propre compte."))
        return redirect('core:users_list')

    if not user_obj.is_active and not user_obj.has_usable_password():
        messages.error(request, _("{username} est en attente d'invitation. Renvoyez l'invitation pour qu'il active son compte lui-même.").format(username=user_obj.username))
        return redirect('core:users_list')

    user_obj.is_active = not user_obj.is_active
    user_obj.save()

    status = "activé" if user_obj.is_active else "désactivé"
    messages.success(request, _("L'utilisateur {username} a été {status}.").format(username=user_obj.username, status=status))
    
    return redirect('core:users_list')


@login_required
def user_change_password(request, user_id):
    """L'admin ne peut plus modifier les mots de passe des autres utilisateurs."""
    messages.error(
        request,
        _("Pour des raisons de sécurité, vous ne pouvez pas modifier le mot de passe "
          "d'un autre utilisateur. Utilisez « Renvoyer l'invitation » pour lui permettre "
          "de définir lui-même son mot de passe.")
    )
    return redirect('core:users_list')


def account_activate(request, uidb64, token):
    """Activation du compte via le lien reçu par email (définition du mot de passe)."""
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    from django.utils.encoding import force_str
    from django.contrib.auth import login as auth_login
    from django.contrib.auth.models import User as DjangoUser
    from django.contrib.auth.forms import SetPasswordForm

    if request.user.is_authenticated:
        return redirect('core:dashboard')

    user = None
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = DjangoUser.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, DjangoUser.DoesNotExist):
        pass

    token_valid = (
        user is not None
        and not user.is_active
        and not user.has_usable_password()
        and default_token_generator.check_token(user, token)
    )

    if not token_valid:
        # Compte déjà activé → rediriger vers login avec message
        if user is not None and user.is_active:
            messages.info(request, _("Ce compte est déjà activé. Connectez-vous normalement."))
            return redirect('core:login')
        return render(request, 'core/users/activate.html', {'valid': False})

    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)
        if form.is_valid():
            form.save()
            user.is_active = True
            user.save(update_fields=['is_active'])
            UserActivity.objects.create(
                user=user,
                action='activate',
                description="Activation du compte via lien email",
            )
            from django.contrib.auth import update_session_auth_hash
            auth_login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            update_session_auth_hash(request, user)
            messages.success(request, _("Bienvenue {name} ! Votre compte est activé.").format(name=user.get_full_name() or user.username))
            return redirect('core:dashboard')
    else:
        form = SetPasswordForm(user)

    return render(request, 'core/users/activate.html', {
        'valid': True,
        'form': form,
        'username': user.username,
        'full_name': user.get_full_name(),
    })


def request_new_activation(request):
    """Page publique : l'utilisateur entre son email pour recevoir un nouveau lien d'activation."""
    from django.contrib.auth.models import User as DjangoUser

    sent = False
    form_error = None

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        if not email:
            form_error = "Veuillez saisir votre adresse email."
        else:
            try:
                user = DjangoUser.objects.get(email__iexact=email, is_active=False)
            except DjangoUser.DoesNotExist:
                user = None

            # Envoyer seulement si le compte est réellement en attente
            if user is not None and not user.has_usable_password():
                activation_link = _build_activation_link(request, user)
                from .notifications import notify_account_invitation
                notify_account_invitation(user, activation_link, invited_by=None)

            # Réponse neutre (ne pas révéler si l'email existe)
            sent = True

    return render(request, 'core/users/request_activation.html', {
        'sent': sent,
        'form_error': form_error,
        'email_value': request.POST.get('email', '') if not sent else '',
    })


@login_required
def resend_invitation(request, user_id):
    """Renvoie l'email d'invitation à un utilisateur inactif (POST uniquement)."""
    from django.contrib.auth.models import User as DjangoUser

    if not request.user.profile.has_manage_users_permission():
        return redirect('core:dashboard')

    if request.method != 'POST':
        return redirect('core:users_list')

    user_obj = get_object_or_404(DjangoUser, pk=user_id)

    if user_obj.is_active:
        messages.warning(request, _("Le compte de {username} est déjà actif.").format(username=user_obj.username))
        return redirect('core:users_list')

    if user_obj.has_usable_password():
        messages.warning(request, _("Le compte de {username} n'est pas en attente d'invitation (il a déjà un mot de passe).").format(username=user_obj.username))
        return redirect('core:users_list')

    if not user_obj.email:
        messages.error(request, _("Ce compte n'a pas d'adresse email."))
        return redirect('core:users_list')

    activation_link = _build_activation_link(request, user_obj)
    invited_by = request.user.get_full_name() or request.user.username
    from .notifications import notify_account_invitation
    ok, msg = notify_account_invitation(user_obj, activation_link, invited_by)

    if ok:
        messages.success(request, _("Invitation renvoyée à {email}.").format(email=user_obj.email))
    else:
        messages.error(request, _("Échec de l'envoi : {msg}").format(msg=msg))

    return redirect('core:users_list')


@login_required
def user_activities(request):
    """Journal des activités"""
    from .models import UserActivity
    
    if not request.user.profile.has_manage_users_permission():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:dashboard')
    
    activities = UserActivity.objects.select_related('user').all()[:100]
    
    context = {
        'activities': activities,
    }
    return render(request, 'core/users/activities.html', context)


def get_client_ip(request):
    """Récupérer l'adresse IP du client"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# ==================== PROJECT CRUD ====================


@login_required
def project_detail(request, project_id):
    """Détail d'un projet"""
    project = get_object_or_404(
        Project.objects.select_related('direction').prefetch_related('milestones', 'needs', 'comments', 'members', 'project_documents'),
        pk=project_id,
    )

    if not request.user.profile.can_view_project(project):
        messages.error(request, _("Vous n'avez pas accès à ce projet."))
        return redirect('core:projects')

    profile = request.user.profile
    can_manage_members = profile.can_manage_project_members(project)
    can_perform_actions = profile.can_perform_actions_as_member(project)
    is_readonly = profile.is_project_readonly(project)
    can_edit_project = profile.can_edit_project(project)
    can_add_milestones = profile.can_add_project_milestones(project)
    can_add_documents = profile.can_add_project_documents(project)
    user_employee = profile.employee

    # Statistiques de charge par membre : {employee_id: {'total': n, 'done': n}}
    from django.db.models import Count, Q as Qm
    milestones_qs = project.milestones.prefetch_related('assigned_to', 'sub_milestones__assigned_to').all()
    member_task_stats = {}
    for m in milestones_qs:
        for emp in m.assigned_to.all():
            s = member_task_stats.setdefault(emp.id, {'total': 0, 'done': 0})
            s['total'] += 1
            if m.completed:
                s['done'] += 1
        for sub in m.sub_milestones.all():
            for emp in sub.assigned_to.all():
                s = member_task_stats.setdefault(emp.id, {'total': 0, 'done': 0})
                s['total'] += 1
                if sub.completed:
                    s['done'] += 1

    # Build member cards for inline quick-assign (only project members)
    members_cards = []
    for pm in project.members.select_related('employee', 'employee__direction').filter(employee__isnull=False):
        emp = pm.employee
        parts = emp.name.split()
        initials = parts[0][0].upper() if parts else '?'
        if len(parts) > 1:
            initials += parts[-1][0].upper()
        task_info = member_task_stats.get(emp.id, {'total': 0, 'done': 0})
        members_cards.append({
            'id': emp.id,
            'name': emp.name,
            'job_role': emp.role,
            'project_role': pm.project_role.name if pm.project_role else '—',
            'direction': emp.direction.code if emp.direction else '',
            'task_count': task_info['total'],
            'task_done': task_info['done'],
            'task_remaining': task_info['total'] - task_info['done'],
            'initials': initials,
            'color': _AVATAR_COLORS[sum(ord(c) for c in emp.name) % len(_AVATAR_COLORS)],
        })

    # Dashboard stats
    from datetime import date, timedelta
    _today = date.today()
    _ml = list(milestones_qs)
    _sc = {}
    _en_retard = 0
    for _m in _ml:
        _key = 'termine' if _m.completed else _m.status
        _sc[_key] = _sc.get(_key, 0) + 1
        if not _m.completed and _m.due_date and _m.due_date < _today:
            _en_retard += 1
    dashboard_stats = {
        'total': len(_ml),
        'termine': _sc.get('termine', 0),
        'en_cours': _sc.get('en_cours', 0),
        'bloque': _sc.get('bloque', 0),
        'en_revision': _sc.get('en_revision', 0),
        'a_faire': _sc.get('a_faire', 0),
        'en_retard': _en_retard,
    }
    overdue_milestones = [m for m in _ml if not m.completed and m.due_date and m.due_date < _today]
    upcoming_milestones = sorted(
        [m for m in _ml if not m.completed and m.due_date and _today <= m.due_date <= _today + timedelta(days=14)],
        key=lambda m: m.due_date,
    )
    open_needs_count = project.needs.filter(status='ouvert').count()

    context = {
        'project': project,
        'milestones': milestones_qs,
        'needs': project.needs.all(),
        'comments': project.comments.all(),
        'activities': project.activities.all()[:20],
        'members': project.members.select_related('employee').all(),
        'documents': project.project_documents.all(),
        'can_manage_members': can_manage_members,
        'can_perform_actions': can_perform_actions,
        'is_readonly': is_readonly,
        'can_edit_project': can_edit_project,
        'can_add_milestones': can_add_milestones,
        'can_add_documents': can_add_documents,
        'user_employee': user_employee,
        'member_task_stats': member_task_stats,
        'members_cards': members_cards,
        'dashboard_stats': dashboard_stats,
        'overdue_milestones': overdue_milestones,
        'upcoming_milestones': upcoming_milestones,
        'open_needs_count': open_needs_count,
    }
    return render(request, 'core/project_detail.html', context)


@login_required
def export_project_activities(request, project_id):
    """Exporter le journal d'activité d'un projet en PDF"""
    project = get_object_or_404(Project, pk=project_id)
    if not request.user.profile.can_view_project(project):
        messages.error(request, _("Vous n'avez pas accès à ce projet."))
        return redirect('core:projects')
    activities = project.activities.all()
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    from urllib.parse import quote
    _pdf_name = f"journal_activite_{project.name.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = (
        f"attachment; filename=\"{_pdf_name}\"; filename*=UTF-8''{quote(_pdf_name)}"
    )
    
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    
    # Style personnalisé pour le titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.grey,
        spaceAfter=30
    )
    
    elements = []
    
    # Titre
    elements.append(Paragraph(f"Journal d'activité", title_style))
    elements.append(Paragraph(f"Projet : {project.name}", subtitle_style))
    elements.append(Paragraph(f"Exporté le {timezone.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 20))
    
    if activities:
        # Tableau des activités
        data = [['Date', 'Utilisateur', 'Action', 'Description']]
        
        for activity in activities:
            data.append([
                activity.created_at.strftime('%d/%m/%Y %H:%M'),
                activity.user,
                activity.get_action_display(),
                activity.description[:50] + '...' if len(activity.description) > 50 else activity.description
            ])
        
        table = Table(data, colWidths=[80, 90, 100, 200])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("Aucune activité enregistrée pour ce projet.", styles['Normal']))
    
    doc.build(elements)
    return response


@login_required
def project_need_create(request, project_id):
    """Créer un besoin sur un projet"""
    from .forms_project import ProjectNeedForm

    project = get_object_or_404(Project, pk=project_id)

    if not request.user.profile.can_add_project_needs(project):
        messages.error(request, _("Vous n'avez pas les permissions pour ajouter des besoins à ce projet."))
        return redirect('core:project_detail', project_id=project.id)

    if request.method != 'POST':
        return redirect('core:project_detail', project_id=project.id)

    form = ProjectNeedForm(request.POST)
    if form.is_valid():
        need = form.save(commit=False)
        need.project = project
        need.created_by = request.user.get_full_name() or request.user.username
        need.save()
        log_project_activity(project, 'ajout_besoin', f"Ajout du besoin '{need.title}'", request.user)
        push_section_refresh(project.id, 'panel-besoins', request.user.get_full_name() or request.user.username)
        messages.success(request, _("Besoin ajouté avec succès."))
    else:
        messages.error(request, _("Impossible d'ajouter le besoin. Vérifiez le formulaire."))

    return redirect('core:project_detail', project_id=project.id)


@login_required
def project_comment_create(request, project_id):
    """Créer un commentaire sur un projet"""
    from .forms_project import ProjectCommentForm

    project = get_object_or_404(Project, pk=project_id)

    if not request.user.profile.can_add_project_comments(project):
        messages.error(request, _("Vous n'avez pas les permissions pour ajouter des commentaires à ce projet."))
        return redirect('core:project_detail', project_id=project.id)

    if request.method != 'POST':
        return redirect('core:project_detail', project_id=project.id)

    form = ProjectCommentForm(request.POST)
    if form.is_valid():
        comment = form.save(commit=False)
        comment.project = project
        comment.created_by = request.user.get_full_name() or request.user.username
        comment.save()
        log_project_activity(project, 'ajout_commentaire', f"Ajout d'un commentaire", request.user)
        from .notifs import notify_commentaire
        notify_commentaire(project, request.user)
        push_section_refresh(project.id, 'panel-commentaires', request.user.get_full_name() or request.user.username)
        messages.success(request, _("Commentaire ajouté."))
    else:
        messages.error(request, _("Impossible d'ajouter le commentaire. Vérifiez le formulaire."))

    return redirect('core:project_detail', project_id=project.id)

@login_required
def project_import(request):
    """Import de projets via Excel → JSON éditable → création en base."""
    if not request.user.profile.can_create_projects():
        messages.error(request, _("Vous n'avez pas les permissions pour créer des projets."))
        return redirect('core:projects')

    directions = Direction.objects.all()
    today = _date.today()
    end_date_default = _date(today.year + 1, today.month, today.day)
    upload_ctx = {
        'step': 'upload',
        'directions': directions,
        'today': today,
        'end_date_default': end_date_default,
    }

    if request.method == 'POST':
        step = request.POST.get('step', 'upload')

        # ── Étape 1 : lecture Excel → JSON affiché dans textarea ──────────
        if step == 'upload':
            uploaded = request.FILES.get('file')
            if not uploaded:
                messages.error(request, _("Veuillez sélectionner un fichier Excel."))
                return render(request, 'core/project_import.html', upload_ctx)

            direction_id   = request.POST.get('direction', '')
            start_date_str = request.POST.get('start_date', str(today))
            end_date_str   = request.POST.get('end_date', str(end_date_default))

            projects_data, error = _excel_to_projects_json(
                uploaded, start_date_str, end_date_str
            )
            if error:
                messages.error(request, f"Erreur de lecture : {error}")
                return render(request, 'core/project_import.html', upload_ctx)

            import json as _json
            json_text = _json.dumps(projects_data, ensure_ascii=False, indent=2)

            direction_obj = Direction.objects.filter(pk=direction_id).first() if direction_id else None
            return render(request, 'core/project_import.html', {
                'step': 'json_preview',
                'json_text': json_text,
                'direction': direction_obj,
                'direction_id': direction_id,
                'directions': directions,
                'count': len(projects_data),
            })

        # ── Étape 2 : import depuis le JSON (édité ou non) ────────────────
        elif step == 'import_json':
            import json as _json
            json_text    = request.POST.get('json_text', '')
            direction_id = request.POST.get('direction_id', '')
            direction    = Direction.objects.filter(pk=direction_id).first() if direction_id else None

            try:
                projects_data = _json.loads(json_text)
            except _json.JSONDecodeError as e:
                messages.error(request, f"JSON invalide : {e}")
                direction_obj = Direction.objects.filter(pk=direction_id).first() if direction_id else None
                return render(request, 'core/project_import.html', {
                    'step': 'json_preview',
                    'json_text': json_text,
                    'direction': direction_obj,
                    'direction_id': direction_id,
                    'directions': directions,
                    'count': 0,
                })

            existing_lower = {n.lower() for n in Project.objects.values_list('name', flat=True)}
            created_count = skipped_count = 0

            for item in projects_data:
                name = str(item.get('name', '')).strip()
                if not name:
                    continue
                if name.lower() in existing_lower:
                    skipped_count += 1
                    continue

                # Dates
                try:
                    start_date = datetime.strptime(item.get('start_date', ''), '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    start_date = today
                try:
                    end_date = datetime.strptime(item.get('end_date', ''), '%Y-%m-%d').date()
                    if end_date <= start_date:
                        raise ValueError
                except (ValueError, TypeError):
                    end_date = _date(start_date.year + 1, start_date.month, start_date.day)

                VALID_PRIORITIES = {'basse', 'moyenne', 'haute'}
                priority = item.get('priority', 'moyenne')
                if priority not in VALID_PRIORITIES:
                    priority = 'moyenne'

                VALID_STATUSES = {'planifie', 'en_cours', 'suspendu', 'termine'}
                status = item.get('status', 'planifie')
                if status not in VALID_STATUSES:
                    status = 'planifie'

                project = Project.objects.create(
                    name=name,
                    manager=str(item.get('manager', ''))[:100],
                    description=item.get('description', ''),
                    priority=priority,
                    status=status,
                    direction=direction,
                    start_date=start_date,
                    end_date=end_date,
                )
                existing_lower.add(name.lower())

                # Jalons
                for order, ms in enumerate(item.get('milestones', []), start=1):
                    ms_name = str(ms.get('name', '')).strip()
                    if not ms_name:
                        continue
                    try:
                        ms_due = datetime.strptime(ms.get('due_date', ''), '%Y-%m-%d').date()
                        if ms_due < start_date or ms_due > end_date:
                            ms_due = end_date
                    except (ValueError, TypeError):
                        ms_due = end_date

                    VALID_MS_STATUSES = {'a_faire', 'en_cours', 'bloque', 'en_revision', 'termine'}
                    ms_status = ms.get('status', 'a_faire')
                    if ms_status not in VALID_MS_STATUSES:
                        ms_status = 'a_faire'
                    ms_completed = ms_status == 'termine'

                    milestone = Milestone.objects.create(
                        project=project,
                        name=ms_name,
                        due_date=ms_due,
                        status=ms_status,
                        completed=ms_completed,
                        order=order,
                    )
                    # Sous-étapes
                    for sub_order, sub_name in enumerate(ms.get('sub_milestones', []), start=1):
                        sub_name = str(sub_name).strip()
                        if sub_name:
                            SubMilestone.objects.create(
                                milestone=milestone,
                                name=sub_name,
                                order=sub_order,
                            )

                log_project_activity(
                    project, 'creation',
                    _("Import JSON : création du projet '%(name)s'") % {'name': project.name},
                    request.user,
                )
                created_count += 1

            if created_count:
                messages.success(request,
                    _("%(n)d projet(s) importé(s).%(s)s") % {
                        'n': created_count,
                        's': f" {skipped_count} doublon(s) ignoré(s)." if skipped_count else '',
                    })
            else:
                messages.warning(request, _("Aucun projet créé — tous les noms existent déjà."))
            return redirect('core:projects')

    return render(request, 'core/project_import.html', upload_ctx)


def _excel_to_projects_json(file_obj, default_start='', default_end=''):
    """
    Lit un .xlsx et retourne (list[dict], error_str).

    Types de feuilles gérés :
      A_multi  : col0=ProjetID rempli → chaque ligne = Milestone, col6=Etapes → SubMilestones
      A_single : col0 vide, col6=Etapes → Milestones (pas de sous-jalons)
      B        : col0=Description de Projet, col5=Etapes → Milestones
      C        : texte libre → lignes = Milestones
      None     : vide/template → projet sans jalons
    """
    import openpyxl, re
    from datetime import datetime as _dt, date as _d

    PRIORITY_MAP = {'low': 'basse', 'medium': 'moyenne', 'high': 'haute'}

    def _fmt_date(val):
        if isinstance(val, (_dt, _d)):
            return val.strftime('%Y-%m-%d')
        if val:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    return _dt.strptime(str(val).strip(), fmt).strftime('%Y-%m-%d')
                except ValueError:
                    pass
        return ''

    def _cell(row, idx, default=''):
        val = row[idx] if len(row) > idx else None
        return str(val).strip() if val is not None else default

    def _map_evo_status(text):
        t = str(text).lower()
        if any(k in t for k in ['complet', 'termin', 'done', 'fini']):
            return 'termine'
        if 'cours' in t:
            return 'en_cours'
        if 'bloqu' in t:
            return 'bloque'
        return 'a_faire'

    def _parse_evolution(evo_text):
        """Retourne dict {num_str: status} depuis le texte Evolution multi-ligne."""
        result = {}
        for line in str(evo_text or '').split('\n'):
            line = line.strip()
            m = re.match(
                r'^(?:Etape|Activit[eé]|Output|R[eé]sultat)[^\d]*(\d+[\.\d]*)\s*[:/]\s*(.+)$',
                line, re.IGNORECASE
            )
            if m:
                result[m.group(1)] = _map_evo_status(m.group(2))
        return result

    def _split_etapes(text, evo_text=''):
        """
        Découpe un bloc multi-lignes d'étapes en liste de dicts {name, status}.
        Gère : Etape N:, Activité N:, Résultat N:/Output N:, bullet points.
        """
        if not text:
            return []
        evo = _parse_evolution(evo_text)
        steps = []
        current_num = None
        for line in str(text).split('\n'):
            line = line.strip().lstrip('•–—·\t ')
            if not line:
                continue
            m = re.match(
                r'^(?:Etape|Activit[eé]|R[eé]sultat\s*\d*\s*/?\s*Output\s*\d*|Output)\s*(\d+[\.\d]*)\s*[:/]?\s*(.*)$',
                line, re.IGNORECASE
            )
            if m:
                current_num = m.group(1)
                content = m.group(2).strip()
                if content:
                    steps.append({'name': content, 'status': evo.get(current_num, 'a_faire')})
            elif line:
                steps.append({'name': line, 'status': evo.get(current_num, 'a_faire') if current_num else 'a_faire'})
        return steps

    def _find_sheet(wb, project_name):
        import difflib as _dl
        name_l = project_name.strip().lower()
        # Correspondance exacte (insensible à la casse)
        for sn in wb.sheetnames:
            if sn.strip().lower() == name_l:
                return wb[sn]
        # Correspondance partielle — nom tronqué à 31 chars ou contenu dans l'autre
        for sn in wb.sheetnames:
            sn_l = sn.strip().lower()
            if name_l.startswith(sn_l) or sn_l.startswith(name_l[:20]):
                return wb[sn]
        # Correspondance approchée (gère les fautes de frappe type Huawey→Huawei)
        candidates = [sn.strip().lower() for sn in wb.sheetnames]
        close = _dl.get_close_matches(name_l, candidates, n=1, cutoff=0.82)
        if close:
            idx = candidates.index(close[0])
            return wb[wb.sheetnames[idx]]
        return None

    def _sheet_type(ws):
        """
        Retourne ('A_multi', cols) | ('A_single', cols) | ('B', cols) | ('C',) | (None,).
        cols = dict mapping champ → index de colonne.
        """
        all_rows = list(ws.iter_rows(max_row=4, values_only=True))

        # Chercher la première ligne non vide (header potentiel)
        header_row = None
        header_idx = 0
        for i, row in enumerate(all_rows):
            if any(v for v in row if v):
                header_row = row
                header_idx = i
                break

        if header_row is None:
            return (None,)

        h0 = str(header_row[0] or '').lower().strip()
        header_str = ' '.join(str(v or '').lower() for v in header_row)
        is_structured = any(k in header_str for k in ['debut', 'etape', 'evolution', 'fin'])

        if not is_structured:
            # Feuille texte libre (Type C) — données dès la première ligne non vide
            return ('C',)

        # Type A : "Projet ID" en col 0
        if 'projet' in h0 and 'id' in h0:
            cols = {'id': 0, 'desc': 1, 'lead': 2, 'debut': 4, 'fin': 5, 'etapes': 6, 'evo': 7}
            data_rows = [
                r for r in all_rows[header_idx + 1:]
                if any(v for v in r if v)
            ]
            if not data_rows:
                return (None,)
            has_named = any(
                r[0] and not isinstance(r[0], (int, float))
                and str(r[0]).strip() not in ('', 'None')
                for r in data_rows
            )
            if has_named:
                return ('A_multi', cols)
            # Col 0 vide ou None → vérifier si Etapes a du contenu
            for dr in data_rows:
                etapes_val = dr[cols['etapes']] if len(dr) > cols['etapes'] else None
                if etapes_val and _split_etapes(str(etapes_val)):
                    return ('A_single', cols)
            return (None,)

        # Type B : "Description" en col 0
        if 'description' in h0:
            cols = {'desc': 0, 'lead': 1, 'debut': 3, 'fin': 4, 'etapes': 5, 'evo': 6}
            return ('B', cols)

        return ('C',)

    # ── Chargement ───────────────────────────────────────────────────────

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        return None, str(e)

    if 'Sommaire' not in wb.sheetnames:
        return None, "La feuille 'Sommaire' est introuvable dans le fichier."

    ws_sommaire = wb['Sommaire']
    projects = []

    for raw in ws_sommaire.iter_rows(min_row=2, values_only=True):
        name = raw[0] if len(raw) > 0 else None
        if not name or not str(name).strip():
            continue

        name     = str(name).strip()
        manager  = _cell(raw, 1)
        desc     = _cell(raw, 2)
        prio_raw = _cell(raw, 3).lower()
        priority = PRIORITY_MAP.get(prio_raw, 'moyenne')

        proj = {
            'name': name,
            'manager': manager,
            'description': desc,
            'priority': priority,
            'status': 'planifie',
            'start_date': default_start,
            'end_date': default_end,
            'milestones': [],
        }

        ws_proj = _find_sheet(wb, name)
        if not ws_proj:
            projects.append(proj)
            continue

        stype = _sheet_type(ws_proj)

        if stype[0] == 'A_multi':
            # Chaque ligne de données = un Milestone ; Etapes col → SubMilestones
            cols = stype[1]
            for row in ws_proj.iter_rows(min_row=2, values_only=True):
                raw_id = row[cols['id']] if len(row) > cols['id'] else None
                if not raw_id or isinstance(raw_id, (int, float)):
                    continue
                ms_name = str(raw_id).strip()
                if not ms_name or ms_name.isdigit() or ms_name.lower().startswith('tache'):
                    continue
                ms_start = _fmt_date(row[cols['debut']] if len(row) > cols['debut'] else None)
                ms_end   = _fmt_date(row[cols['fin']]   if len(row) > cols['fin']   else None)
                evo_text = row[cols['evo']]    if len(row) > cols['evo']    else None
                et_text  = row[cols['etapes']] if len(row) > cols['etapes'] else None
                sub_steps = _split_etapes(et_text, evo_text)

                # Statut global du milestone : si l'évolution global est lisible en col evo
                ms_status = _map_evo_status(str(evo_text)) if evo_text else 'a_faire'

                proj['milestones'].append({
                    'name': ms_name,
                    'due_date': ms_end or proj['end_date'],
                    'status': ms_status,
                    'sub_milestones': [s['name'] for s in sub_steps],
                })
                if not proj['start_date'] and ms_start:
                    proj['start_date'] = ms_start
                if not proj['end_date'] and ms_end:
                    proj['end_date'] = ms_end

        elif stype[0] == 'A_single':
            # Une seule ligne de données ; Etapes col → Milestones (plats)
            cols = stype[1]
            for row in ws_proj.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v):
                    continue
                ms_start = _fmt_date(row[cols['debut']] if len(row) > cols['debut'] else None)
                ms_end   = _fmt_date(row[cols['fin']]   if len(row) > cols['fin']   else None)
                rich_desc = _cell(row, cols['desc'])
                lead      = _cell(row, cols['lead']) or manager
                evo_text  = row[cols['evo']]    if len(row) > cols['evo']    else None
                et_text   = row[cols['etapes']] if len(row) > cols['etapes'] else None
                steps     = _split_etapes(et_text, evo_text)

                if ms_start:
                    proj['start_date'] = ms_start
                if ms_end:
                    proj['end_date'] = ms_end
                if rich_desc and not proj['description']:
                    proj['description'] = rich_desc
                if lead:
                    proj['manager'] = lead

                for s in steps:
                    proj['milestones'].append({
                        'name': s['name'],
                        'due_date': ms_end or proj['end_date'],
                        'status': s['status'],
                        'sub_milestones': [],
                    })
                break  # une seule ligne de données

        elif stype[0] == 'B':
            # Une ligne ; col0=Description, col5=Etapes → Milestones (plats)
            cols = stype[1]
            for row in ws_proj.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row if v):
                    continue
                rich_desc = _cell(row, cols['desc'])
                lead      = _cell(row, cols['lead']) or manager
                ms_start  = _fmt_date(row[cols['debut']] if len(row) > cols['debut'] else None)
                ms_end    = _fmt_date(row[cols['fin']]   if len(row) > cols['fin']   else None)
                evo_text  = row[cols['evo']]    if len(row) > cols['evo']    else None
                et_text   = row[cols['etapes']] if len(row) > cols['etapes'] else None
                steps     = _split_etapes(et_text, evo_text)

                if ms_start:
                    proj['start_date'] = ms_start
                if ms_end:
                    proj['end_date'] = ms_end
                if rich_desc:
                    proj['description'] = rich_desc
                if lead:
                    proj['manager'] = lead

                for s in steps:
                    proj['milestones'].append({
                        'name': s['name'],
                        'due_date': ms_end or proj['end_date'],
                        'status': s['status'],
                        'sub_milestones': [],
                    })
                break

        elif stype[0] == 'C':
            # Feuille texte libre : chaque cellule non vide = un Milestone
            for row in ws_proj.iter_rows(min_row=1, values_only=True):
                for cell in row:
                    val = str(cell).strip() if cell else ''
                    if val and val not in ('None', ''):
                        proj['milestones'].append({
                            'name': val,
                            'due_date': proj['end_date'],
                            'status': 'a_faire',
                            'sub_milestones': [],
                        })

        projects.append(proj)

    return projects, None


@login_required
def project_create(request):
    """Créer un nouveau projet"""
    from .forms_project import ProjectForm
    
    if not request.user.profile.can_create_projects():
        messages.error(request, _("Vous n'avez pas les permissions pour créer un projet."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, user=request.user)
        if form.is_valid():
            project = form.save(commit=False)
            # Sync manager string from manager_employee FK
            if project.manager_employee:
                project.manager = project.manager_employee.name
            project.save()
            form.save_m2m()
            # Le responsable choisi doit faire partie de l'équipe projet
            _ensure_manager_in_team(project, request.user)
            log_project_activity(project, 'creation', f"Création du projet '{project.name}'", request.user)
            messages.success(request, _("Projet créé avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = ProjectForm(user=request.user)

    return render(request, 'core/project_form.html', {'form': form, 'title': 'Nouveau projet'})


@login_required
def project_edit(request, project_id):
    """Modifier un projet"""
    from .forms_project import ProjectForm
    
    project = get_object_or_404(Project, pk=project_id)

    # Vérifier les permissions
    if not request.user.profile.can_edit_project(project):
        messages.error(request, _("Vous n'avez pas les permissions pour modifier ce projet."))
        return redirect('core:projects')

    # Migration silencieuse : sync manager_employee depuis le string manager si FK est vide
    if project.manager and not project.manager_employee_id:
        emp = Employee.objects.filter(name__iexact=project.manager.strip(), is_external=False).first()
        if emp:
            project.manager_employee = emp
            project.save(update_fields=['manager_employee'])

    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            old_manager_id = project.manager_employee_id
            project = form.save(commit=False)
            # Sync manager string from manager_employee FK
            if project.manager_employee:
                project.manager = project.manager_employee.name
            project.save()
            form.save_m2m()
            # Si le responsable a changé, l'ajouter automatiquement à l'équipe
            if project.manager_employee_id != old_manager_id:
                _ensure_manager_in_team(project, request.user)
            log_project_activity(project, 'modification', "Modification des informations du projet", request.user)
            push_project_meta(project)
            messages.success(request, _("Projet modifié avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = ProjectForm(instance=project)

    return render(request, 'core/project_form.html', {'form': form, 'project': project, 'title': f'Modifier {project.name}'})


@login_required
def project_delete(request, project_id):
    """Supprimer un projet"""
    # Permission delete:Project (RBAC) ou DG/admin en fallback
    if not (request.user.profile.can('delete', 'Project') or request.user.profile.is_directeur_general()):
        messages.error(request, _("Vous n'avez pas la permission de supprimer ce projet."))
        return redirect('core:projects')
    
    project = get_object_or_404(Project, pk=project_id)
    
    if request.method == 'POST':
        project.delete()
        messages.success(request, _("Projet supprimé."))
        return redirect('core:projects')
    
    return render(request, 'core/confirm_delete.html', {'object': project, 'type': 'projet', 'back_url': 'core:projects'})


# ==================== MILESTONE CRUD ====================

@login_required
def milestone_detail(request, milestone_id):
    """Vue détail d'un jalon"""
    milestone = get_object_or_404(
        Milestone.objects.select_related('project', 'project__direction', 'assigned_by')
                         .prefetch_related('assigned_to', 'sub_milestones__assigned_to'),
        pk=milestone_id,
    )
    project = milestone.project

    if not request.user.profile.can_view_project(project):
        messages.error(request, _("Vous n'avez pas accès à ce projet."))
        return redirect('core:projects')

    profile = request.user.profile
    can_edit    = profile.can_edit_project_milestones(project)
    user_employee = profile.employee

    sub_milestones = milestone.sub_milestones.prefetch_related('assigned_to').all()
    sub_total      = sub_milestones.count()
    sub_done       = sub_milestones.filter(completed=True).count()

    return render(request, 'core/milestone_detail.html', {
        'milestone':      milestone,
        'project':        project,
        'can_edit':       can_edit,
        'user_employee':  user_employee,
        'sub_milestones': sub_milestones,
        'sub_total':      sub_total,
        'sub_done':       sub_done,
    })


@login_required
def milestone_create(request, project_id):
    """Créer un jalon pour un projet"""
    from .forms_project import MilestoneForm
    
    project = get_object_or_404(Project, pk=project_id)
    
    # Permission check - utiliser la nouvelle méthode
    if not request.user.profile.can_add_project_milestones(project):
        messages.error(request, _("Vous n'avez pas les permissions pour ajouter des jalons à ce projet."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        form = MilestoneForm(project=project, data=request.POST)
        if form.is_valid():
            milestone = form.save(commit=False)
            milestone.project = project
            milestone.save()
            assigned_ids = request.POST.getlist('assigned_to')
            assigned_emps = list(Employee.objects.filter(pk__in=assigned_ids)) if assigned_ids else []
            if assigned_emps:
                milestone.assigned_to.set(assigned_emps)
                milestone.assigned_by = request.user
                milestone.save(update_fields=['assigned_by'])
                from .notifications import notify_assignment
                assigned_by_name = request.user.get_full_name() or request.user.username
                for emp in assigned_emps:
                    success, msg = notify_assignment(emp, 'jalon', milestone.name, project.name, assigned_by_name, project=project, due_date=milestone.due_date)
                    if success:
                        messages.info(request, msg)
            log_project_activity(project, 'ajout_jalon', f"Ajout du jalon '{milestone.name}'", request.user)
            project.refresh_from_db()
            push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
            messages.success(request, _("Jalon créé avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = MilestoneForm(project=project)

    project_members, other_employees = _get_project_member_cards(project)
    return render(request, 'core/milestone_form.html', {
        'form': form, 'project': project, 'title': 'Nouveau jalon',
        'project_members': project_members, 'other_employees': other_employees,
        'initial_assignee_ids': [],
    })


@login_required
def milestone_edit(request, milestone_id):
    """Modifier un jalon"""
    from .forms_project import MilestoneForm
    
    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project
    
    # Permission check
    if not request.user.profile.can_edit_project_milestones(project):
        messages.error(request, _("Vous n'avez pas les permissions pour modifier ce jalon."))
        return redirect('core:projects')
    
    old_assigned_ids = set(milestone.assigned_to.values_list('id', flat=True))
    was_completed = milestone.completed
    if request.method == 'POST':
        form = MilestoneForm(project=project, data=request.POST, instance=milestone)
        if form.is_valid():
            milestone = form.save(commit=False)
            milestone.save()
            assigned_ids = request.POST.getlist('assigned_to')
            assigned_emps = list(Employee.objects.filter(pk__in=assigned_ids)) if assigned_ids else []
            milestone.assigned_to.set(assigned_emps)
            new_assigned_ids = {emp.id for emp in assigned_emps}
            newly_added_ids = new_assigned_ids - old_assigned_ids
            if newly_added_ids:
                milestone.assigned_by = request.user
                milestone.save(update_fields=['assigned_by'])
                from .notifications import notify_assignment
                assigned_by_name = request.user.get_full_name() or request.user.username
                for emp in assigned_emps:
                    if emp.id in newly_added_ids:
                        success, msg = notify_assignment(emp, 'jalon', milestone.name, project.name, assigned_by_name, project=project, due_date=milestone.due_date)
                        if success:
                            messages.info(request, msg)
            if milestone.completed and not was_completed:
                from .notifications import notify_task_completed
                for emp in assigned_emps:
                    notify_task_completed('jalon', milestone.name, project, emp, milestone.assigned_by, request.user)
            log_project_activity(project, 'modif_jalon', f"Modification du jalon '{milestone.name}'", request.user)
            project.refresh_from_db()
            push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
            messages.success(request, _("Jalon modifié avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = MilestoneForm(project=project, instance=milestone)

    project_members, other_employees = _get_project_member_cards(project)
    initial_assignee_ids = list(old_assigned_ids)
    return render(request, 'core/milestone_form.html', {
        'form': form, 'project': project, 'milestone': milestone,
        'title': f'Modifier {milestone.name}',
        'project_members': project_members, 'other_employees': other_employees,
        'initial_assignee_ids': initial_assignee_ids,
    })


@login_required
def milestone_delete(request, milestone_id):
    """Supprimer un jalon"""
    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project
    
    # Permission check
    if not request.user.profile.can_edit_project_milestones(project):
        messages.error(request, _("Vous n'avez pas les permissions pour supprimer ce jalon."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        milestone_name = milestone.name
        milestone.delete()
        log_project_activity(project, 'suppr_jalon', f"Suppression du jalon '{milestone_name}'", request.user)
        project.refresh_from_db()
        push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
        messages.success(request, _("Jalon supprimé."))
        return redirect('core:project_detail', project_id=project.id)
    
    return render(request, 'core/confirm_delete.html', {'object': milestone, 'type': 'jalon', 'back_url': 'core:project_detail', 'back_args': {'project_id': project.id}})


# ==================== SUB-MILESTONE CRUD ====================

@login_required
def sub_milestone_create(request, milestone_id):
    """Créer une sous-étape pour un jalon"""
    from .forms_project import SubMilestoneForm
    from .models import Milestone
    
    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project
    
    # Permission check
    if not request.user.profile.can_add_project_milestones(project):
        messages.error(request, _("Vous n'avez pas les permissions pour ajouter des sous-étapes."))
        return redirect('core:project_detail', project_id=project.id)
    
    if request.method == 'POST':
        form = SubMilestoneForm(milestone=milestone, data=request.POST)
        if form.is_valid():
            sub_milestone = form.save(commit=False)
            sub_milestone.milestone = milestone
            sub_milestone.save()
            assigned_ids = request.POST.getlist('assigned_to')
            assigned_emps = list(Employee.objects.filter(pk__in=assigned_ids)) if assigned_ids else []
            if assigned_emps:
                sub_milestone.assigned_to.set(assigned_emps)
                sub_milestone.assigned_by = request.user
                sub_milestone.save(update_fields=['assigned_by'])
                from .notifications import notify_assignment
                assigned_by_name = request.user.get_full_name() or request.user.username
                for emp in assigned_emps:
                    success, msg = notify_assignment(emp, 'sous-étape', sub_milestone.name, project.name, assigned_by_name, project=project, due_date=sub_milestone.due_date)
                    if success:
                        messages.info(request, msg)
            log_project_activity(project, 'ajout_sous_etape', f"Ajout de la sous-étape '{sub_milestone.name}' au jalon '{milestone.name}'", request.user)
            project.refresh_from_db()
            push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
            messages.success(request, _("Sous-étape ajoutée avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = SubMilestoneForm(milestone=milestone)

    project_members, other_employees = _get_project_member_cards(project)
    return render(request, 'core/sub_milestone_form.html', {
        'form': form, 'milestone': milestone, 'project': project,
        'title': f'Nouvelle sous-étape pour "{milestone.name}"',
        'project_members': project_members, 'other_employees': other_employees,
        'initial_assignee_ids': [],
    })


@login_required
def sub_milestone_edit(request, sub_milestone_id):
    """Modifier une sous-étape"""
    from .forms_project import SubMilestoneForm
    from .models import SubMilestone
    
    sub_milestone = get_object_or_404(SubMilestone.objects.select_related('milestone__project'), pk=sub_milestone_id)
    milestone = sub_milestone.milestone
    project = milestone.project
    
    # Permission check
    if not request.user.profile.can_edit_project_milestones(project):
        messages.error(request, _("Vous n'avez pas les permissions pour modifier cette sous-étape."))
        return redirect('core:project_detail', project_id=project.id)
    
    old_assigned_ids = set(sub_milestone.assigned_to.values_list('id', flat=True))
    was_completed = sub_milestone.completed
    if request.method == 'POST':
        form = SubMilestoneForm(milestone=milestone, data=request.POST, instance=sub_milestone)
        if form.is_valid():
            sub_milestone = form.save(commit=False)
            sub_milestone.save()
            assigned_ids = request.POST.getlist('assigned_to')
            assigned_emps = list(Employee.objects.filter(pk__in=assigned_ids)) if assigned_ids else []
            sub_milestone.assigned_to.set(assigned_emps)
            new_assigned_ids = {emp.id for emp in assigned_emps}
            newly_added_ids = new_assigned_ids - old_assigned_ids
            if newly_added_ids:
                sub_milestone.assigned_by = request.user
                sub_milestone.save(update_fields=['assigned_by'])
                from .notifications import notify_assignment
                assigned_by_name = request.user.get_full_name() or request.user.username
                for emp in assigned_emps:
                    if emp.id in newly_added_ids:
                        success, msg = notify_assignment(emp, 'sous-étape', sub_milestone.name, project.name, assigned_by_name, project=project, due_date=sub_milestone.due_date)
                        if success:
                            messages.info(request, msg)
            if sub_milestone.completed and not was_completed:
                from .notifications import notify_task_completed
                for emp in assigned_emps:
                    notify_task_completed('sous-étape', sub_milestone.name, project, emp, sub_milestone.assigned_by, request.user)
            log_project_activity(project, 'modif_sous_etape', f"Modification de la sous-étape '{sub_milestone.name}'", request.user)
            project.refresh_from_db()
            push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
            messages.success(request, _("Sous-étape modifiée avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = SubMilestoneForm(milestone=milestone, instance=sub_milestone)

    project_members, other_employees = _get_project_member_cards(project)
    initial_assignee_ids = list(old_assigned_ids)
    return render(request, 'core/sub_milestone_form.html', {
        'form': form, 'milestone': milestone, 'project': project,
        'sub_milestone': sub_milestone, 'title': f'Modifier "{sub_milestone.name}"',
        'project_members': project_members, 'other_employees': other_employees,
        'initial_assignee_ids': initial_assignee_ids,
    })


@login_required
def sub_milestone_delete(request, sub_milestone_id):
    """Supprimer une sous-étape"""
    from .models import SubMilestone
    
    sub_milestone = get_object_or_404(SubMilestone.objects.select_related('milestone__project'), pk=sub_milestone_id)
    milestone = sub_milestone.milestone
    project = milestone.project
    
    # Permission check
    if not request.user.profile.can_edit_project_milestones(project):
        messages.error(request, _("Vous n'avez pas les permissions pour supprimer cette sous-étape."))
        return redirect('core:project_detail', project_id=project.id)
    
    if request.method == 'POST':
        sub_name = sub_milestone.name
        sub_milestone.delete()
        log_project_activity(project, 'suppr_sous_etape', f"Suppression de la sous-étape '{sub_name}'", request.user)
        project.refresh_from_db()
        push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
        messages.success(request, _("Sous-étape supprimée."))
        return redirect('core:project_detail', project_id=project.id)
    
    return render(request, 'core/confirm_delete.html', {
        'object': sub_milestone, 
        'type': 'sous-étape', 
        'back_url': 'core:project_detail', 
        'back_args': {'project_id': project.id}
    })


@login_required
def milestone_quick_assign(request, milestone_id):
    """AJAX toggle : ajouter/retirer un employé des responsables d'un jalon.
    POST assigned_to=<id> pour toggler, assigned_to= (vide) pour tout effacer."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project
    if not request.user.profile.can_edit_project_milestones(project):
        return JsonResponse({'ok': False, 'error': 'Permission refusée'}, status=403)

    emp_id = request.POST.get('assigned_to') or None

    if emp_id:
        try:
            emp = Employee.objects.get(pk=emp_id)
        except Employee.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Employé introuvable'}, status=400)
        if milestone.assigned_to.filter(pk=emp_id).exists():
            milestone.assigned_to.remove(emp)
        else:
            milestone.assigned_to.add(emp)
            milestone.assigned_by = request.user
            milestone.save(update_fields=['assigned_by'])
            from .notifications import notify_assignment
            notify_assignment(
                emp, 'jalon', milestone.name, project.name,
                request.user.get_full_name() or request.user.username,
                project=project, due_date=milestone.due_date,
            )
    else:
        milestone.assigned_to.clear()

    log_project_activity(project, 'modif_jalon', f"Réassignation du jalon '{milestone.name}'", request.user)
    assignees = list(milestone.assigned_to.values('id', 'name'))
    push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username)
    return JsonResponse({
        'ok': True,
        'assigned_names': [a['name'] for a in assignees],
        'assigned_ids': [a['id'] for a in assignees],
    })


@login_required
def sub_milestone_quick_assign(request, sub_milestone_id):
    """AJAX toggle : ajouter/retirer un employé des responsables d'une sous-étape."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    sub = get_object_or_404(SubMilestone.objects.select_related('milestone__project'), pk=sub_milestone_id)
    project = sub.milestone.project
    if not request.user.profile.can_edit_project_milestones(project):
        return JsonResponse({'ok': False, 'error': 'Permission refusée'}, status=403)

    emp_id = request.POST.get('assigned_to') or None

    if emp_id:
        try:
            emp = Employee.objects.get(pk=emp_id)
        except Employee.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Employé introuvable'}, status=400)
        if sub.assigned_to.filter(pk=emp_id).exists():
            sub.assigned_to.remove(emp)
        else:
            sub.assigned_to.add(emp)
            sub.assigned_by = request.user
            sub.save(update_fields=['assigned_by'])
            from .notifications import notify_assignment
            notify_assignment(
                emp, 'sous-étape', sub.name, project.name,
                request.user.get_full_name() or request.user.username,
                project=project, due_date=sub.due_date,
            )
    else:
        sub.assigned_to.clear()

    log_project_activity(project, 'modif_sous_etape', f"Réassignation de la sous-étape '{sub.name}'", request.user)
    assignees = list(sub.assigned_to.values('id', 'name'))
    push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username)
    return JsonResponse({
        'ok': True,
        'assigned_names': [a['name'] for a in assignees],
        'assigned_ids': [a['id'] for a in assignees],
    })


@login_required
def sub_milestone_toggle(request, sub_milestone_id):
    """Basculer le statut complété d'une sous-étape.
    Autorisé : utilisateurs avec can_add_project_milestones OU le responsable assigné."""
    from .models import SubMilestone

    sub_milestone = get_object_or_404(SubMilestone.objects.select_related('milestone__project'), pk=sub_milestone_id)
    milestone = sub_milestone.milestone
    project = milestone.project

    can_toggle = (
        request.user.profile.can_edit_project_milestones(project)
        or _user_is_assignee(request.user, sub_milestone.assigned_to)
    )
    if not can_toggle:
        messages.error(request, _("Vous n'avez pas les permissions pour modifier cette sous-étape."))
        return redirect('core:project_detail', project_id=project.id)

    was_completed = sub_milestone.completed
    sub_milestone.completed = not sub_milestone.completed
    sub_milestone.save()

    # Refresh pour obtenir les valeurs recalculées par les signaux post_save
    milestone.refresh_from_db()
    project.refresh_from_db()

    status = "complétée" if sub_milestone.completed else "non complétée"
    log_project_activity(project, 'toggle_sous_etape', f"Sous-étape '{sub_milestone.name}' marquée comme {status}", request.user)
    if sub_milestone.completed and not was_completed:
        from .notifications import notify_task_completed
        from .notifs import notify_sous_etape_completed
        notify_sous_etape_completed(sub_milestone, request.user)
        for emp in sub_milestone.assigned_to.all():
            success, msg = notify_task_completed('sous-étape', sub_milestone.name, project, emp, sub_milestone.assigned_by, request.user)
            if not success:
                messages.warning(request, _("Sous-étape terminée mais notification email échouée : {msg}").format(msg=msg))

    from .notifs import push_project_update
    push_project_update(project.id, {
        'event': 'sous_etape_toggle',
        'sub_id': sub_milestone.id,
        'milestone_id': milestone.id,
        'completed': sub_milestone.completed,
        'milestone_progress': milestone.progress,
        'milestone_completed': milestone.completed,
        'project_progress': project.progress,
        'actor': request.user.get_full_name() or request.user.username,
    })

    messages.success(request, _("Sous-étape marquée comme {status}.").format(status=status))
    return redirect('core:project_detail', project_id=project.id)


@login_required
def milestone_toggle(request, milestone_id):
    """Basculer le statut complété d'un jalon (sans sous-étapes).
    Autorisé : utilisateurs avec can_add_project_milestones OU le responsable assigné."""
    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project

    can_toggle = (
        request.user.profile.can_edit_project_milestones(project)
        or _user_is_assignee(request.user, milestone.assigned_to)
    )
    if not can_toggle:
        messages.error(request, _("Vous n'avez pas les permissions pour modifier ce jalon."))
        return redirect('core:project_detail', project_id=project.id)

    if milestone.sub_milestones.exists():
        messages.warning(request, _("Ce jalon contient des sous-étapes : sa progression est calculée automatiquement."))
        return redirect('core:project_detail', project_id=project.id)

    from django.utils import timezone
    was_completed = milestone.completed
    milestone.completed = not milestone.completed
    milestone.manual_progress = 100 if milestone.completed else 0
    milestone.status = 'termine' if milestone.completed else 'a_faire'
    milestone.completed_at = timezone.now() if milestone.completed else None
    milestone.save()

    # Refresh pour obtenir la progression projet recalculée par les signaux post_save
    project.refresh_from_db()

    label = "complété" if milestone.completed else "non complété"
    log_project_activity(project, 'toggle_jalon', f"Jalon '{milestone.name}' marqué comme {label}", request.user)
    if milestone.completed and not was_completed:
        from .notifications import notify_task_completed
        from .notifs import notify_jalon_completed
        notify_jalon_completed(milestone, request.user)
        for emp in milestone.assigned_to.all():
            success, msg = notify_task_completed('jalon', milestone.name, project, emp, milestone.assigned_by, request.user)
            if not success:
                messages.warning(request, _("Jalon terminé mais notification email échouée : {msg}").format(msg=msg))

    from .notifs import push_project_update
    push_project_update(project.id, {
        'event': 'milestone_toggle',
        'milestone_id': milestone.id,
        'completed': milestone.completed,
        'progress': milestone.progress,
        'project_progress': project.progress,
        'actor': request.user.get_full_name() or request.user.username,
    })

    messages.success(request, _("Jalon marqué comme {label}.").format(label=label))
    return redirect('core:project_detail', project_id=project.id)


@login_required
def milestone_update_status(request, milestone_id):
    """Changer le statut d'un jalon (AJAX-compatible, méthode POST)."""
    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project
    if not (request.user.profile.can_edit_project_milestones(project) or _user_is_assignee(request.user, milestone.assigned_to)):
        messages.error(request, _("Permissions insuffisantes."))
        return redirect('core:project_detail', project_id=project.id)

    new_status = request.POST.get('status', '')
    valid = dict(Milestone.TASK_STATUS_CHOICES)
    if new_status not in valid:
        messages.error(request, _("Statut invalide."))
        return redirect('core:project_detail', project_id=project.id)

    from django.utils import timezone
    old_status = milestone.status
    milestone.status = new_status
    if new_status == 'termine':
        milestone.completed = True
        milestone.manual_progress = 100
        if not milestone.completed_at:
            milestone.completed_at = timezone.now()
    else:
        milestone.completed = False
        if new_status == 'a_faire':
            milestone.manual_progress = 0
        milestone.completed_at = None
    milestone.save()

    log_project_activity(project, 'modif_statut_jalon', f"Statut de '{milestone.name}' : {valid.get(old_status, old_status)} → {valid[new_status]}", request.user)
    if milestone.completed and old_status != 'termine':
        from .notifications import notify_task_completed
        for emp in milestone.assigned_to.all():
            notify_task_completed('jalon', milestone.name, project, emp, milestone.assigned_by, request.user)
    project.refresh_from_db()
    push_milestone_status(project.id, milestone.id, new_status, valid[new_status], milestone.completed, milestone.progress, project.progress)
    messages.success(request, _("Statut mis à jour : {status}.").format(status=valid[new_status]))
    return redirect('core:project_detail', project_id=project.id)


@login_required
def project_need_update_status(request, need_id):
    """Mettre à jour le statut d'un besoin (résolu, rejeté, etc.)."""
    from .models import ProjectNeed
    from django.utils import timezone
    need = get_object_or_404(ProjectNeed.objects.select_related('project'), pk=need_id)
    project = need.project

    if not request.user.profile.can_view_project(project):
        messages.error(request, _("Accès refusé."))
        return redirect('core:projects')
    if not (request.user.profile.can_add_project_needs(project) or request.user.profile.can_manage_project_members(project)):
        messages.error(request, _("Permissions insuffisantes."))
        return redirect('core:project_detail', project_id=project.id)

    new_status = request.POST.get('status', '')
    valid = dict(ProjectNeed.NEED_STATUS_CHOICES)
    if new_status not in valid:
        messages.error(request, _("Statut invalide."))
        return redirect('core:project_detail', project_id=project.id)

    need.status = new_status
    if new_status in ('resolu', 'rejete'):
        need.resolved_by = request.user.get_full_name() or request.user.username
        need.resolved_at = timezone.now()
    else:
        need.resolved_by = ''
        need.resolved_at = None
    need.save()
    log_project_activity(
        project, 'modification',
        f"Statut du besoin '{need.title}' → {valid[new_status]}",
        request.user,
    )
    push_section_refresh(project.id, 'panel-besoins', request.user.get_full_name() or request.user.username)
    messages.success(request, _("Besoin marqué : {status}.").format(status=valid[new_status]))
    return redirect('core:project_detail', project_id=project.id)


@login_required
def my_tasks(request):
    """Vue 'Mes tâches' : toutes les tâches assignées à l'utilisateur courant."""
    from .models import Milestone, SubMilestone
    from collections import defaultdict

    user_employee = getattr(request.user.profile, 'employee', None)
    if not user_employee:
        messages.info(request, _("Votre compte n'est pas encore lié à un employé. Contactez un administrateur."))
        return redirect('core:dashboard')

    milestones_qs = (
        Milestone.objects
        .filter(assigned_to=user_employee)
        .select_related('project__direction')
        .prefetch_related('sub_milestones')
        .order_by('completed', 'due_date')
    )
    sub_milestones_qs = (
        SubMilestone.objects
        .filter(assigned_to=user_employee)
        .select_related('milestone__project__direction', 'milestone')
        .order_by('completed', 'due_date')
    )

    by_project = defaultdict(lambda: {'project': None, '_items': [], 'total': 0, 'done': 0})

    for m in milestones_qs:
        can_toggle = (m.sub_milestones.count() == 0)
        entry = by_project[m.project_id]
        entry['project'] = m.project
        entry['_items'].append({
            'id': m.id, 'type': 'jalon', 'name': m.name,
            'milestone_name': None, 'due_date': m.due_date,
            'completed': m.completed, 'completed_at': m.completed_at,
            'status': m.status, 'can_toggle': can_toggle,
        })
        entry['total'] += 1
        if m.completed:
            entry['done'] += 1

    for s in sub_milestones_qs:
        p = s.milestone.project
        entry = by_project[p.id]
        entry['project'] = p
        entry['_items'].append({
            'id': s.id, 'type': 'sous_etape', 'name': s.name,
            'milestone_name': s.milestone.name, 'due_date': s.due_date,
            'completed': s.completed, 'completed_at': s.completed_at,
            'status': None, 'can_toggle': True,
        })
        entry['total'] += 1
        if s.completed:
            entry['done'] += 1

    for entry in by_project.values():
        entry['milestones'] = sorted(
            entry.pop('_items'),
            key=lambda x: (x['completed'], str(x['due_date']) if x['due_date'] else '9999-99-99')
        )

    task_groups = sorted(by_project.values(), key=lambda e: e['project'].name)

    from datetime import date as _date
    today = _date.today()
    total_tasks = sum(e['total'] for e in by_project.values())
    tasks_done  = sum(e['done']  for e in by_project.values())
    tasks_overdue = sum(
        1 for e in by_project.values()
        for item in e['milestones']
        if not item['completed'] and item['due_date'] and item['due_date'] < today
    )

    return render(request, 'core/my_tasks.html', {
        'task_groups': task_groups,
        'user_employee': user_employee,
        'task_stats': {
            'total': total_tasks,
            'done': tasks_done,
            'pending': total_tasks - tasks_done - tasks_overdue,
            'overdue': tasks_overdue,
        },
    })


@login_required
def api_project_task_create(request, project_id):
    """Créer une tâche de projet depuis le board Planner."""
    import json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    project = get_object_or_404(Project, pk=project_id)
    if not request.user.profile.can_add_project_milestones(project):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body or '{}')
        title = (data.get('title') or '').strip()
        status = data.get('status') or 'a_faire'
        if not title:
            return JsonResponse({'error': 'Le titre est requis.'}, status=400)

        max_order = project.milestones.aggregate(models.Max('order'))['order__max'] or 0
        progress = 0
        completed = False
        if status == 'en_cours':
            progress = 50
        elif status == 'termine':
            progress = 100
            completed = True

        milestone = Milestone.objects.create(
            project=project,
            name=title,
            manual_progress=progress,
            completed=completed,
            order=max_order + 1,
        )
        project.recalculate_progress()
        project.refresh_from_db()
        log_project_activity(project, 'ajout_jalon', f"Ajout de la tâche '{milestone.name}'", request.user)
        push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
        return JsonResponse({
            'success': True,
            'task': {
                'id': milestone.id,
                'name': milestone.name,
                'progress': milestone.progress,
                'completed': milestone.completed,
                'sub_count': 0,
                'edit_url': reverse('core:milestone_edit', args=[milestone.id]),
                'sub_url': reverse('core:sub_milestone_create', args=[milestone.id]),
                'update_url': reverse('core:api_project_task_update', args=[milestone.id]),
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_project_task_update(request, milestone_id):
    """Mettre à jour une tâche depuis la fiche latérale Planner."""
    import json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project
    if not request.user.profile.can_edit_project_milestones(project):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body or '{}')
        if 'title' in data:
            title = (data.get('title') or '').strip()
            if title:
                milestone.name = title
        if 'description' in data:
            milestone.need = data.get('description') or ''
        if 'status' in data:
            status = data.get('status')
            valid_statuses = [s[0] for s in Milestone.TASK_STATUS_CHOICES]
            if status in valid_statuses:
                milestone.status = status
                if status == 'a_faire':
                    milestone.manual_progress = 0
                    milestone.completed = False
                elif status == 'en_cours':
                    if milestone.manual_progress == 0:
                        milestone.manual_progress = 50
                    milestone.completed = False
                elif status == 'termine':
                    milestone.manual_progress = 100
                    milestone.completed = True
                else:
                    milestone.completed = False
        if 'progress' in data:
            progress = max(0, min(100, int(data.get('progress') or 0)))
            milestone.manual_progress = progress
            milestone.completed = progress >= 100

        milestone.save()
        project.recalculate_progress()
        project.refresh_from_db()
        log_project_activity(project, 'modif_jalon', f"Mise à jour de la tâche '{milestone.name}'", request.user)
        push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username, project_progress=project.progress)
        return JsonResponse({'success': True, 'task': {'id': milestone.id, 'name': milestone.name, 'progress': milestone.progress, 'completed': milestone.completed, 'update_url': reverse('core:api_project_task_update', args=[milestone.id])}})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ==================== PROJECT FOLDER & DOCUMENT CRUD ====================

@login_required
def project_folder_create(request, project_id):
    """Créer un dossier pour un projet"""
    from .forms_project import ProjectFolderForm

    project = get_object_or_404(Project, pk=project_id)
    parent_id = request.GET.get('parent')
    parent_folder = None
    if parent_id:
        parent_folder = get_object_or_404(ProjectFolder, pk=parent_id, project=project)

    if not request.user.profile.can_add_project_documents(project):
        messages.error(request, _("Vous n'avez pas les permissions pour gérer les dossiers de ce projet."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        form = ProjectFolderForm(project, request.POST)
        if form.is_valid():
            folder = form.save(commit=False)
            folder.project = project
            folder.save()
            log_project_activity(project, 'ajout_dossier', f"Création du dossier '{folder.name}'", request.user)
            push_section_refresh(project.id, 'panel-documents', request.user.get_full_name() or request.user.username)
            messages.success(request, _("Dossier créé avec succès."))
            if folder.parent_id:
                return redirect('core:project_folder_detail', folder_id=folder.id)
            return redirect('core:project_detail', project_id=project.id)
    else:
        if parent_folder:
            form = ProjectFolderForm(project, initial={'parent': parent_folder})
        else:
            form = ProjectFolderForm(project)
    
    return render(request, 'core/project_folder_form.html', {'form': form, 'project': project, 'title': 'Nouveau dossier'})


@login_required
def project_folder_detail(request, folder_id):
    """Détail d'un dossier de projet (sous-dossiers + documents)"""
    folder = get_object_or_404(ProjectFolder.objects.select_related('project', 'parent'), pk=folder_id)
    project = folder.project

    if not request.user.profile.can_view_project(project):
        messages.error(request, _("Vous n'avez pas accès à ce projet."))
        return redirect('core:projects')

    subfolders = folder.subfolders.all().order_by('name')
    documents = folder.documents.all().order_by('-uploaded_at')

    context = {
        'project': project,
        'folder': folder,
        'subfolders': subfolders,
        'documents': documents,
    }
    return render(request, 'core/project_folder_detail.html', context)


@login_required
def project_folder_edit(request, folder_id):
    """Modifier un dossier de projet"""
    from .forms_project import ProjectFolderForm

    folder = get_object_or_404(ProjectFolder.objects.select_related('project'), pk=folder_id)
    project = folder.project

    if not request.user.profile.can_add_project_documents(project):
        messages.error(request, _("Vous n'avez pas les permissions pour gérer les dossiers de ce projet."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        form = ProjectFolderForm(project, request.POST, instance=folder)
        if form.is_valid():
            form.save()
            log_project_activity(project, 'modification', f"Modification du dossier '{folder.name}'", request.user)
            push_section_refresh(project.id, 'panel-documents', request.user.get_full_name() or request.user.username)
            messages.success(request, _("Dossier modifié avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = ProjectFolderForm(project, instance=folder)
    
    return render(request, 'core/project_folder_form.html', {'form': form, 'project': project, 'folder': folder, 'title': f'Modifier {folder.name}'})


@login_required
def project_folder_delete(request, folder_id):
    """Supprimer un dossier de projet"""
    folder = get_object_or_404(ProjectFolder.objects.select_related('project'), pk=folder_id)
    project = folder.project

    if not request.user.profile.can_edit_project_documents(project):
        messages.error(request, _("Vous n'avez pas les permissions pour gérer les dossiers de ce projet."))
        return redirect('core:projects')

    if request.method == 'POST':
        folder_name = folder.name
        folder.delete()
        log_project_activity(project, 'suppr_dossier', f"Suppression du dossier '{folder_name}'", request.user)
        push_section_refresh(project.id, 'panel-documents', request.user.get_full_name() or request.user.username)
        messages.success(request, _("Dossier supprimé."))
        return redirect('core:project_detail', project_id=project.id)
    
    return render(request, 'core/confirm_delete.html', {'object': folder, 'type': 'dossier', 'back_url': 'core:project_detail', 'back_args': {'project_id': project.id}})


@login_required
def project_document_create(request, project_id):
    """Créer un document pour un projet"""
    from .forms_project import ProjectDocumentForm

    project = get_object_or_404(Project, pk=project_id)
    folder_id = request.GET.get('folder')
    initial_folder = None
    if folder_id:
        initial_folder = get_object_or_404(ProjectFolder, pk=folder_id, project=project)

    if not request.user.profile.can_add_project_documents(project):
        messages.error(request, _("Vous n'avez pas les permissions pour ajouter des documents à ce projet."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        form = ProjectDocumentForm(project, request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.project = project
            doc.uploaded_by = request.user.get_full_name() or request.user.username
            doc.save()
            log_project_activity(project, 'ajout_document', f"Ajout du document '{doc.title}'", request.user)
            push_section_refresh(project.id, 'panel-documents', request.user.get_full_name() or request.user.username)
            messages.success(request, _("Document ajouté avec succès."))
            if doc.folder_id:
                return redirect('core:project_folder_detail', folder_id=doc.folder_id)
            return redirect('core:project_detail', project_id=project.id)
    else:
        if initial_folder:
            form = ProjectDocumentForm(project, initial={'folder': initial_folder})
        else:
            form = ProjectDocumentForm(project)
    
    return render(request, 'core/project_document_form.html', {'form': form, 'project': project, 'title': 'Nouveau document'})


@login_required
def project_document_edit(request, doc_id):
    """Modifier un document de projet"""
    from .forms_project import ProjectDocumentForm

    doc = get_object_or_404(ProjectDocument.objects.select_related('project'), pk=doc_id)
    project = doc.project

    if not request.user.profile.can_edit_project_documents(project):
        messages.error(request, _("Vous n'avez pas les permissions pour modifier les documents de ce projet."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        form = ProjectDocumentForm(project, request.POST, request.FILES, instance=doc)
        if form.is_valid():
            form.save()
            log_project_activity(project, 'ajout_document', f"Modification du document '{doc.title}'", request.user)
            push_section_refresh(project.id, 'panel-documents', request.user.get_full_name() or request.user.username)
            messages.success(request, _("Document modifié avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = ProjectDocumentForm(project, instance=doc)
    
    return render(request, 'core/project_document_form.html', {'form': form, 'project': project, 'doc': doc, 'title': f'Modifier {doc.title}'})


@login_required
def project_document_delete(request, doc_id):
    """Supprimer un document de projet"""
    doc = get_object_or_404(ProjectDocument.objects.select_related('project'), pk=doc_id)
    project = doc.project

    if not request.user.profile.can_edit_project_documents(project):
        messages.error(request, _("Vous n'avez pas les permissions pour supprimer les documents de ce projet."))
        return redirect('core:projects')
    
    if request.method == 'POST':
        doc_title = doc.title
        doc.delete()
        log_project_activity(project, 'suppr_document', f"Suppression du document '{doc_title}'", request.user)
        push_section_refresh(project.id, 'panel-documents', request.user.get_full_name() or request.user.username)
        messages.success(request, _("Document supprimé."))
        return redirect('core:project_detail', project_id=project.id)

    return render(request, 'core/confirm_delete.html', {'object': doc, 'type': 'document', 'back_url': 'core:project_detail', 'back_args': {'project_id': project.id}})


@login_required
def project_document_download(request, doc_id):
    """Télécharger un document de projet — redirect avec fl_attachment pour forcer le téléchargement."""
    from django.http import HttpResponseRedirect

    doc = get_object_or_404(ProjectDocument.objects.select_related('project'), pk=doc_id)
    project = doc.project

    if not request.user.profile.can_view_project(project):
        messages.error(request, _("Vous n'avez pas accès à ce projet."))
        return redirect('core:projects')
    
    if not doc.file:
        messages.error(request, _("Fichier non trouvé."))
        return redirect('core:project_detail', project_id=project.id)

    url = doc.file.replace('/upload/', '/upload/fl_attachment/')
    return HttpResponseRedirect(url)


@login_required
def project_document_preview(request, doc_id):
    """Prévisualiser un document de projet (PDF, image...)."""
    import mimetypes
    from urllib.parse import urlparse, unquote

    doc = get_object_or_404(ProjectDocument.objects.select_related('project'), pk=doc_id)
    project = doc.project

    if not request.user.profile.can_view_project(project):
        messages.error(request, _("Vous n'avez pas accès à ce projet."))
        return redirect('core:projects')

    if not doc.file:
        messages.error(request, _("Aucun fichier attaché à ce document."))
        return redirect('core:project_detail', project_id=project.id)

    file_url = reverse('core:project_document_file_proxy', args=[doc.id])
    path_part = urlparse(doc.file).path
    url_filename = unquote(path_part.rsplit('/', 1)[-1]) if '/' in path_part else unquote(path_part)
    ext = url_filename.rsplit('.', 1)[-1].lower() if '.' in url_filename else ''
    filename = f"{doc.title}.{ext}" if ext else doc.title

    mime_type, _encoding = mimetypes.guess_type(url_filename)
    mime_type = mime_type or 'application/octet-stream'

    if mime_type.startswith('image/'):
        preview_type = 'image'
    elif mime_type == 'application/pdf':
        preview_type = 'pdf'
    else:
        preview_type = 'unsupported'

    if doc.folder_id:
        back_url = reverse('core:project_folder_detail', args=[doc.folder_id])
    else:
        back_url = reverse('core:project_detail', args=[project.id])

    context = {
        'document': doc,
        'project': project,
        'file_url': file_url,
        'mime_type': mime_type,
        'preview_type': preview_type,
        'filename': filename,
        'back_url': back_url,
        'download_url': reverse('core:project_document_download', args=[doc.id]),
    }
    return render(request, 'core/document_preview.html', context)


# ==================== PROJECT MEMBER CRUD ====================

@login_required
def project_member_add(request, project_id):
    """Ajouter un membre à un projet"""
    from .forms_project import ProjectMemberForm
    
    project = get_object_or_404(Project, pk=project_id)

    # Permission check - utiliser la nouvelle méthode
    if not request.user.profile.can_add_project_members(project):
        messages.error(request, _("Vous n'avez pas les permissions pour ajouter des membres à ce projet."))
        return redirect('core:project_detail', project_id=project.id)
    
    directions = Direction.objects.all()

    # Calculate available employees count
    existing_member_ids = project.members.values_list('employee_id', flat=True)
    available_employees_count = Employee.objects.exclude(id__in=existing_member_ids).count()

    from .models import ProjectRole as PRole
    import json as _json

    def _role_perms(r):
        perms = set(r.permissions.values_list('action', 'subject'))
        return {
            'is_system':           r.is_system,
            'name':                r.name,
            'can_manage_members':  ('manage', 'ProjectMember') in perms,
            'can_edit_project':    ('update', 'Project')       in perms,
            'can_add_milestones':  ('create', 'Milestone')     in perms,
            'can_add_documents':   ('create', 'Document')      in perms,
            'can_add_needs':       ('create', 'Request')       in perms,
            'can_add_comments':    ('create', 'Comment')       in perms,
        }

    project_roles_meta = {
        str(r.pk): _role_perms(r)
        for r in PRole.objects.prefetch_related('permissions').all()
    }

    context = {
        'project': project,
        'title': 'Ajouter un membre',
        'directions': directions,
        'available_employees_count': available_employees_count,
        'role_choices': [(r.slug, r.name) for r in Role.objects.all().order_by('name')],
        'project_roles_meta_json': _json.dumps(project_roles_meta),
    }
    
    if request.method == 'POST':
        member_type = request.POST.get('member_type', 'existing')  # 'existing' | 'new_internal' | 'new_external'

        if member_type in ('new_internal', 'new_external'):
            is_external = (member_type == 'new_external')
            new_name = request.POST.get('new_employee_name', '').strip()
            new_role_val = request.POST.get('new_employee_role', '').strip()
            new_phone = request.POST.get('new_employee_phone', '').strip()
            new_email = request.POST.get('new_employee_email', '').strip()
            new_org = request.POST.get('new_employee_organization', '').strip()
            project_role_id = request.POST.get('project_role')
            project_role_obj = None
            if project_role_id:
                from .models import ProjectRole
                project_role_obj = ProjectRole.objects.filter(pk=project_role_id).first()

            errors = []
            # Validation/normalisation du téléphone
            normalized_phone = new_phone
            if new_phone:
                from core.fields import validate_phone_number
                try:
                    normalized_phone = validate_phone_number(new_phone)
                except Exception as exc:
                    errors.append(str(exc))

            if not new_name:
                errors.append("Le nom est requis.")
            if not new_role_val:
                errors.append("Le poste / fonction est requis.")
            if not is_external and not project.direction_id:
                errors.append("Ce projet n'a pas de direction. Assignez-en une avant d'ajouter un employé interne.")
            if is_external and not new_org:
                errors.append("L'organisation est obligatoire pour une personne externe.")
            if not is_external and not new_email:
                errors.append("L'email est obligatoire pour un employé interne (nécessaire pour la création de compte).")
            if new_email:
                from django.core.validators import validate_email as _validate_email
                from django.core.exceptions import ValidationError as _VE
                try:
                    _validate_email(new_email)
                except _VE:
                    errors.append(_("L'adresse email n'est pas valide."))
                else:
                    if Employee.objects.filter(email__iexact=new_email).exists():
                        errors.append(_("Un employé avec l'email « %(email)s » existe déjà.") % {'email': new_email})

            if errors:
                for e in errors:
                    messages.error(request, e)
                context.update({
                    'member_type': member_type,
                    'new_employee_name': new_name,
                    'new_employee_role': new_role_val,
                    'new_employee_phone': new_phone,
                    'new_employee_email': new_email,
                    'new_employee_organization': new_org,
                    'form': ProjectMemberForm(project, request.POST),
                })
                return render(request, 'core/project_member_form.html', context)

            try:
                from django.db import transaction as _tx
                with _tx.atomic():
                    new_employee = Employee.objects.create(
                        name=new_name,
                        direction=project.direction if not is_external else None,
                        role=new_role_val,
                        phone=normalized_phone,
                        email=new_email,
                        is_external=is_external,
                        organization=new_org if is_external else '',
                    )
                    member = ProjectMember.objects.create(project=project, employee=new_employee, project_role=project_role_obj)
                    log_project_activity(project, 'ajout_membre', f"Ajout du membre '{new_name}' ({project_role_obj.name if project_role_obj else 'sans rôle'})", request.user)

                    # --- Créer un compte de connexion (flow invitation) ---
                    create_user = request.POST.get('create_user') == 'on'
                    if create_user and new_employee.email:
                        from django.contrib.auth.models import User as DjangoUser
                        from .models import Role as SysRole
                        default_role_slug = 'visiteur' if is_external else 'employe'
                        new_user_role_slug = request.POST.get('new_user_role', default_role_slug) or default_role_slug
                        sys_role = SysRole.objects.filter(slug=new_user_role_slug).first()

                        new_username = _generate_username(new_employee.name)
                        name_parts = new_employee.name.split() if new_employee.name else []
                        new_user = DjangoUser.objects.create(
                            username=new_username,
                            email=new_employee.email,
                            first_name=name_parts[0] if name_parts else '',
                            last_name=' '.join(name_parts[1:]) if len(name_parts) > 1 else '',
                            is_active=False,
                        )
                        new_user.set_unusable_password()
                        new_user.save()
                        UserProfile.objects.filter(user=new_user).update(
                            employee=new_employee,
                            role=sys_role,
                            direction=new_employee.direction,
                        )
                        activation_link = _build_activation_link(request, new_user)
                        from .notifications import notify_account_invitation
                        ok, msg = notify_account_invitation(new_user, activation_link, invited_by=request.user.get_full_name() or request.user.username)
                        if ok:
                            messages.success(request, _("Invitation envoyée à {email} (identifiant : {username}).").format(email=new_employee.email, username=new_username))
                        else:
                            messages.warning(request, _("Compte '{username}' créé mais l'email n'a pas pu être envoyé : {msg}.").format(username=new_username, msg=msg))
                    elif create_user and not new_employee.email:
                        messages.warning(request, _("Compte non créé : {name} n'a pas d'adresse email.").format(name=new_employee.name))

                from .notifications import notify_project_member_added
                from .notifs import notify_membre_ajoute
                notify_membre_ajoute(member, request.user)
                success, msg = notify_project_member_added(member, request.user)
                if success:
                    messages.info(request, msg)

                label = "externe" if is_external else "interne"
                push_section_refresh(project.id, 'panel-equipe', request.user.get_full_name() or request.user.username)
                messages.success(request, _("Personne {label} '{name}' créée et ajoutée au projet.").format(label=label, name=new_name))
                return redirect('core:project_detail', project_id=project.id)
            except Exception as exc:
                messages.error(request, _("Erreur : {err}").format(err=exc))
                context.update({'member_type': member_type, 'form': ProjectMemberForm(project, request.POST)})
                return render(request, 'core/project_member_form.html', context)
        else:
            # Employé existant
            form = ProjectMemberForm(project, request.POST)
            if form.is_valid():
                member = form.save(commit=False)
                member.project = project
                member.save()
                log_project_activity(project, 'ajout_membre', f"Ajout du membre '{member.employee.name}' ({member.project_role.name if member.project_role else 'sans rôle'})", request.user)
                from .notifications import notify_project_member_added
                from .notifs import notify_membre_ajoute
                notify_membre_ajoute(member, request.user)
                success, msg = notify_project_member_added(member, request.user)
                if success:
                    messages.info(request, msg)
                push_section_refresh(project.id, 'panel-equipe', request.user.get_full_name() or request.user.username)
                messages.success(request, _("Membre ajouté avec succès."))
                return redirect('core:project_detail', project_id=project.id)
            context['form'] = form
            return render(request, 'core/project_member_form.html', context)
    else:
        form = ProjectMemberForm(project)

    context['form'] = form
    return render(request, 'core/project_member_form.html', context)


@login_required
def project_member_edit(request, member_id):
    """Modifier le rôle d'un membre de projet"""
    from .forms_project import ProjectMemberForm
    
    member = get_object_or_404(ProjectMember.objects.select_related('project', 'employee'), pk=member_id)
    project = member.project

    if not request.user.profile.can_manage_project_members(project):
        messages.error(request, _("Vous n'avez pas les permissions pour modifier les membres de ce projet."))
        return redirect('core:project_detail', project_id=project.id)
    
    if request.method == 'POST':
        form = ProjectMemberForm(project, request.POST, instance=member)
        if form.is_valid():
            form.save()
            new_role_name = member.project_role.name if member.project_role else 'sans rôle'
            log_project_activity(project, 'modification', f"Rôle de '{member.employee.name}' modifié → '{new_role_name}'", request.user)
            push_section_refresh(project.id, 'panel-equipe', request.user.get_full_name() or request.user.username)
            messages.success(request, _("Rôle du membre modifié avec succès."))
            return redirect('core:project_detail', project_id=project.id)
    else:
        form = ProjectMemberForm(project, instance=member)
    
    # Calculate available employees count
    existing_member_ids = project.members.values_list('employee_id', flat=True)
    available_employees_count = Employee.objects.exclude(id__in=existing_member_ids).count()
    
    from .models import ProjectRole as PRole
    import json as _json

    def _role_perms(r):
        perms = set(r.permissions.values_list('action', 'subject'))
        return {
            'is_system':           r.is_system,
            'name':                r.name,
            'can_manage_members':  ('manage', 'ProjectMember') in perms,
            'can_edit_project':    ('update', 'Project')       in perms,
            'can_add_milestones':  ('create', 'Milestone')     in perms,
            'can_add_documents':   ('create', 'Document')      in perms,
            'can_add_needs':       ('create', 'Request')       in perms,
            'can_add_comments':    ('create', 'Comment')       in perms,
        }

    project_roles_meta = {
        str(r.pk): _role_perms(r)
        for r in PRole.objects.prefetch_related('permissions').all()
    }

    return render(request, 'core/project_member_form.html', {
        'form': form,
        'project': project,
        'member': member,
        'title': f'Modifier {member.employee.name}',
        'available_employees_count': available_employees_count,
        'role_choices': [(r.slug, r.name) for r in Role.objects.all().order_by('name')],
        'project_roles_meta_json': _json.dumps(project_roles_meta),
    })


@login_required
def project_member_delete(request, member_id):
    """Supprimer un membre d'un projet"""
    member = get_object_or_404(ProjectMember.objects.select_related('project', 'employee'), pk=member_id)
    project = member.project

    if not request.user.profile.can_manage_project_members(project):
        messages.error(request, _("Vous n'avez pas les permissions pour supprimer des membres de ce projet."))
        return redirect('core:project_detail', project_id=project.id)
    
    if request.method == 'POST':
        member_name = member.employee.name
        member.delete()
        log_project_activity(project, 'retrait_membre', f"Retrait du membre '{member_name}'", request.user)
        push_section_refresh(project.id, 'panel-equipe', request.user.get_full_name() or request.user.username)
        messages.success(request, _("Membre supprimé du projet."))
        return redirect('core:project_detail', project_id=project.id)
    
    return render(request, 'core/confirm_delete.html', {'object': member, 'type': 'membre', 'back_url': 'core:project_detail', 'back_args': {'project_id': project.id}})


# ==================== DOCUMENT CRUD ====================

@login_required
def document_create(request):
    """Créer un nouveau document"""
    from .forms_project import DocumentForm
    
    if not request.user.profile.can_approve_documents():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:documents')
    
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, _("Document créé avec succès."))
            return redirect('core:documents')
    else:
        form = DocumentForm()
    
    return render(request, 'core/document_form.html', {'form': form, 'title': 'Nouveau document'})


@login_required
def document_edit(request, doc_id):
    """Modifier un document"""
    from .forms_project import DocumentForm
    
    if not request.user.profile.can_approve_documents():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:documents')
    
    doc = get_object_or_404(Document, pk=doc_id)
    
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance=doc)
        if form.is_valid():
            form.save()
            messages.success(request, _("Document modifié avec succès."))
            return redirect('core:documents')
    else:
        form = DocumentForm(instance=doc)
    
    return render(request, 'core/document_form.html', {'form': form, 'document': doc, 'title': f'Modifier {doc.title}'})


@login_required
def document_delete(request, doc_id):
    """Supprimer un document"""
    if not request.user.profile.can_approve_documents():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:documents')
    
    doc = get_object_or_404(Document, pk=doc_id)
    
    if request.method == 'POST':
        doc.delete()
        messages.success(request, _("Document supprimé."))
        return redirect('core:documents')
    
    return render(request, 'core/confirm_delete.html', {'object': doc, 'type': 'document', 'back_url': 'core:documents'})


@login_required
def document_sign(request, doc_id):
    """Signer un document"""
    if not request.user.profile.can_approve_documents():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:documents')
    
    doc = get_object_or_404(Document, pk=doc_id)
    doc.status = 'signe'
    doc.signed_at = timezone.now().date()
    doc.save()
    messages.success(request, _("Document '{title}' signé.").format(title=doc.title))
    return redirect('core:documents')


@login_required
def document_validate(request, doc_id):
    """Valider un document"""
    if not request.user.profile.can_approve_documents():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:documents')
    
    doc = get_object_or_404(Document, pk=doc_id)
    doc.status = 'a_signer'
    doc.save()
    messages.success(request, _("Document '{title}' validé, en attente de signature.").format(title=doc.title))
    return redirect('core:documents')


@login_required
def document_file_proxy(request, doc_id):
    """Rediriger vers le fichier du document (fichiers publics Cloudinary)."""
    from django.http import HttpResponseRedirect

    doc = get_object_or_404(Document, pk=doc_id)

    profile = request.user.profile
    direction_id = getattr(profile, 'direction_id', None)
    if not profile.is_directeur_general():
        if not direction_id or doc.direction_id != direction_id:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        if not profile.can_approve_documents():
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

    if not doc.file:
        return HttpResponseRedirect('/')

    return HttpResponseRedirect(doc.file)


@login_required
def project_document_file_proxy(request, doc_id):
    """Rediriger vers le fichier du document de projet (fichiers publics Cloudinary)."""
    from django.http import HttpResponseRedirect

    doc = get_object_or_404(ProjectDocument.objects.select_related('project'), pk=doc_id)
    project = doc.project

    if not request.user.profile.can_view_project(project):
        return HttpResponseRedirect('/')

    if not doc.file:
        return HttpResponseRedirect('/')

    return HttpResponseRedirect(doc.file)


@login_required
def document_download(request, doc_id):
    """Télécharger un document — redirection directe vers Cloudinary."""
    from django.http import HttpResponseRedirect

    doc = get_object_or_404(Document, pk=doc_id)

    profile = request.user.profile
    direction_id = getattr(profile, 'direction_id', None)
    if not profile.is_directeur_general():
        if not direction_id or doc.direction_id != direction_id:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        if not profile.can_approve_documents():
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied

    if not doc.file:
        messages.error(request, _("Aucun fichier attaché à ce document."))
        return redirect('core:documents')

    url = doc.file.replace('/upload/', '/upload/fl_attachment/')
    return HttpResponseRedirect(url)


@login_required
def document_preview(request, doc_id):
    """Prévisualiser un document global (PDF, image...)."""
    import mimetypes
    from urllib.parse import urlparse, unquote

    doc = get_object_or_404(Document, pk=doc_id)

    profile = request.user.profile
    direction_id = getattr(profile, 'direction_id', None)
    if not profile.is_directeur_general():
        if not direction_id or doc.direction_id != direction_id:
            messages.error(request, _("Vous n'avez pas accès à ce document."))
            return redirect('core:documents')
        if not profile.can_approve_documents():
            messages.error(request, _("Vous n'avez pas accès à ce document."))
            return redirect('core:documents')

    if not doc.file:
        messages.error(request, _("Aucun fichier attaché à ce document."))
        return redirect('core:documents')

    file_url = reverse('core:document_file_proxy', args=[doc.id])
    path_part = urlparse(doc.file).path
    url_filename = unquote(path_part.rsplit('/', 1)[-1]) if '/' in path_part else unquote(path_part)
    ext = url_filename.rsplit('.', 1)[-1].lower() if '.' in url_filename else ''
    filename = f"{doc.title}.{ext}" if ext else doc.title

    mime_type, _encoding = mimetypes.guess_type(url_filename)
    mime_type = mime_type or 'application/octet-stream'

    if mime_type.startswith('image/'):
        preview_type = 'image'
    elif mime_type == 'application/pdf':
        preview_type = 'pdf'
    else:
        preview_type = 'unsupported'

    context = {
        'document': doc,
        'file_url': file_url,
        'mime_type': mime_type,
        'preview_type': preview_type,
        'filename': filename,
        'back_url': reverse('core:documents'),
        'download_url': reverse('core:document_download', args=[doc.id]),
    }
    return render(request, 'core/document_preview.html', context)


# ==================== REQUEST CRUD ====================

@login_required
def request_create(request):
    """Créer une nouvelle demande"""
    from .forms_project import RequestForm

    profile = request.user.profile
    # L'utilisateur doit appartenir à une direction OU avoir la permission d'approuver les demandes
    if not profile.direction_id and not profile.can_approve_requests():
        messages.error(request, _("Vous devez appartenir à une direction pour soumettre une demande."))
        return redirect('core:dashboard')

    if request.method == 'POST':
        form = RequestForm(request.POST)
        if form.is_valid():
            req = form.save(commit=False)
            req.status = 'en_attente'
            req.created_by = request.user.get_full_name() or request.user.username
            # Forcer la direction de l'utilisateur si non-admin
            if not profile.can_approve_requests() and profile.direction_id:
                req.direction_id = profile.direction_id
            req.save()
            messages.success(request, _("Demande créée avec succès."))
            return redirect('core:requests')
    else:
        form = RequestForm()

    return render(request, 'core/request_form.html', {'form': form, 'title': 'Nouvelle demande'})


@login_required
def request_approve(request, req_id):
    """Approuver une demande"""
    if not request.user.profile.has_approve_requests_permission():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:requests')
    
    req = get_object_or_404(Request, pk=req_id)
    _ra_profile = request.user.profile
    if not (_ra_profile.is_admin() or _ra_profile.is_directeur_general()):
        if req.direction_id and req.direction_id != _ra_profile.direction_id:
            messages.error(request, _("Vous ne pouvez approuver que les demandes de votre direction."))
            return redirect('core:requests')
    req.status = 'approuve'
    req.approved_at = timezone.now().date()
    req.save()
    messages.success(request, _("Demande '{title}' approuvée.").format(title=req.title))
    return redirect('core:requests')


@login_required
def request_reject(request, req_id):
    """Rejeter une demande"""
    if not request.user.profile.has_approve_requests_permission():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:requests')
    
    req = get_object_or_404(Request, pk=req_id)
    _rr_profile = request.user.profile
    if not (_rr_profile.is_admin() or _rr_profile.is_directeur_general()):
        if req.direction_id and req.direction_id != _rr_profile.direction_id:
            messages.error(request, _("Vous ne pouvez rejeter que les demandes de votre direction."))
            return redirect('core:requests')
    req.status = 'rejete'
    req.save()
    messages.success(request, _("Demande '{title}' rejetée.").format(title=req.title))
    return redirect('core:requests')


# ==================== PARTNER CRUD ====================

@login_required
def partner_create(request):
    """Créer un nouveau partenaire"""
    from .forms_project import PartnerForm
    
    if not request.user.profile.can_manage_partners():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:partners')
    
    if request.method == 'POST':
        form = PartnerForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, _("Partenaire créé avec succès."))
            return redirect('core:partners')
    else:
        form = PartnerForm()
    
    return render(request, 'core/partner_form.html', {'form': form, 'title': 'Nouveau partenaire'})


@login_required
def partner_edit(request, partner_id):
    """Modifier un partenaire"""
    from .forms_project import PartnerForm
    
    if not request.user.profile.can_manage_partners():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:partners')
    
    partner = get_object_or_404(Partner, pk=partner_id)
    
    if request.method == 'POST':
        form = PartnerForm(request.POST, request.FILES, instance=partner)
        if form.is_valid():
            form.save()
            messages.success(request, _("Partenaire modifié avec succès."))
            return redirect('core:partners')
    else:
        form = PartnerForm(instance=partner)
    
    return render(request, 'core/partner_form.html', {'form': form, 'partner': partner, 'title': f'Modifier {partner.name}'})


@login_required
def partner_delete(request, partner_id):
    """Supprimer un partenaire"""
    if not request.user.profile.can_manage_partners():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:partners')
    
    partner = get_object_or_404(Partner, pk=partner_id)
    
    if request.method == 'POST':
        partner.delete()
        messages.success(request, _("Partenaire supprimé."))
        return redirect('core:partners')
    
    return render(request, 'core/confirm_delete.html', {'object': partner, 'type': 'partenaire', 'back_url': 'core:partners'})


def _user_can_manage_employee(user, employee=None):
    """Verifie si l'utilisateur peut gerer (creer/modifier/supprimer) un employe.
    - Admin / DG / superuser : tous
    - Directeur : uniquement les employes de sa direction
    - Rôle avec manage:Employee : uniquement sa direction
    """
    profile = getattr(user, 'profile', None)
    if not profile:
        return False
    if user.is_superuser or profile.is_admin() or profile.is_directeur_general():
        return True
    if profile.can('manage', 'Employee'):
        if employee is None:
            return profile.direction_id is not None
        return profile.direction_id is not None and employee.direction_id == profile.direction_id
    return False


# Employee management views
@login_required
def employee_create(request):
    """Créer un nouvel employé"""
    from .forms_project import EmployeeForm

    if not _user_can_manage_employee(request.user):
        messages.error(request, _("Vous n'avez pas les permissions pour gérer les employés."))
        return redirect('core:resources')

    profile = request.user.profile

    if request.method == 'POST':
        form = EmployeeForm(request.POST, user=request.user)
        if form.is_valid():
            employee = form.save(commit=False)
            # Forcer la direction pour tout rôle non-global avec une direction assignée
            if not (profile.is_admin() or profile.is_directeur_general()) and profile.direction_id:
                employee.direction_id = profile.direction_id
            employee.save()

            system_role = form.cleaned_data.get('system_role')
            if system_role:
                ok, msg = _create_or_update_user_for_employee(request, employee, system_role)
                if ok:
                    messages.success(request, msg)
                else:
                    messages.warning(request, msg)
            else:
                messages.success(request, _("Employé créé avec succès."))
            return redirect('core:resources')
    else:
        form = EmployeeForm(user=request.user)

    return render(request, 'core/employee_form.html', {'form': form, 'title': 'Nouvel employé'})


@login_required
def employee_edit(request, employee_id):
    """Modifier un employé existant"""
    from .forms_project import EmployeeForm

    employee = get_object_or_404(Employee, pk=employee_id)

    if not _user_can_manage_employee(request.user, employee):
        messages.error(request, _("Vous ne pouvez gérer que les employés de votre direction."))
        return redirect('core:resources')

    profile = request.user.profile

    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            if not (profile.is_admin() or profile.is_directeur_general()) and profile.direction_id:
                obj.direction_id = profile.direction_id
            obj.save()

            # Synchroniser User lié si la fiche est modifiée
            try:
                linked_profile = obj.user_profile
                linked_user = linked_profile.user
                name_parts = obj.name.split() if obj.name else []
                linked_user.first_name = name_parts[0] if name_parts else ''
                linked_user.last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                if obj.email:
                    linked_user.email = obj.email
                linked_user.save(update_fields=['first_name', 'last_name', 'email'])
            except (AttributeError, ValueError) as e:
                from .error_logging import ErrorLogger
                ErrorLogger.log_exception(e, context={'function': 'employee_update', 'employee': obj.name}, user=request.user)

            system_role = form.cleaned_data.get('system_role')
            if system_role:
                ok, msg = _create_or_update_user_for_employee(request, obj, system_role)
                if ok:
                    messages.success(request, msg)
                else:
                    messages.warning(request, msg)
            else:
                messages.success(request, _("Employé modifié avec succès."))
            return redirect('core:resources')
    else:
        form = EmployeeForm(instance=employee, user=request.user)

    return render(request, 'core/employee_form.html', {'form': form, 'employee': employee, 'title': f'Modifier {employee.name}'})


@login_required
def employee_delete(request, employee_id):
    """Supprimer un employé"""
    employee = get_object_or_404(Employee, pk=employee_id)

    if not _user_can_manage_employee(request.user, employee):
        messages.error(request, _("Vous ne pouvez supprimer que les employés de votre direction."))
        return redirect('core:resources')

    # Calcul des dépendances pour afficher l'avertissement
    managed_projects = employee.managed_projects.all()
    memberships = employee.project_memberships.select_related('project').all()
    leave_count = employee.leave_requests.count()
    has_account = hasattr(employee, 'user_profile') and employee.user_profile is not None

    if request.method == 'POST':
        if leave_count > 0:
            messages.error(request, _("Impossible de supprimer {name} : {count} demande(s) de congé sont liées à cette fiche.").format(name=employee.name, count=leave_count))
            return redirect('core:resources')
        try:
            name = employee.name
            employee.delete()
            messages.success(request, _("Employé « {name} » supprimé.").format(name=name))
        except Exception as e:
            messages.error(request, _("Suppression impossible : {err}").format(err=e))
        return redirect('core:resources')

    context = {
        'object': employee,
        'type': 'employé',
        'back_url': 'core:resources',
        'managed_projects': managed_projects,
        'memberships': memberships,
        'leave_count': leave_count,
        'has_account': has_account,
    }
    return render(request, 'core/confirm_delete.html', context)


# Budget management views
@login_required
def budget_create(request):
    """Créer un nouveau budget"""
    from .forms_project import BudgetForm
    
    if not request.user.profile.can_manage_budgets():
        messages.error(request, _("Vous n'avez pas les permissions pour gérer les budgets."))
        return redirect('core:resources')
    
    if request.method == 'POST':
        form = BudgetForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, _("Budget créé avec succès."))
            return redirect('core:resources')
    else:
        form = BudgetForm(user=request.user)

    return render(request, 'core/budget_form.html', {'form': form, 'title': 'Nouveau budget'})


@login_required
def budget_edit(request, budget_id):
    """Modifier un budget existant"""
    from .forms_project import BudgetForm
    
    if not request.user.profile.can_manage_budgets():
        messages.error(request, _("Vous n'avez pas les permissions pour gérer les budgets."))
        return redirect('core:resources')
    
    budget = get_object_or_404(Budget, pk=budget_id)

    _be_profile = request.user.profile
    _be_is_global = request.user.is_superuser or _be_profile.is_admin() or _be_profile.is_directeur_general()
    if not _be_is_global:
        _accessible_ids = list(get_accessible_projects_qs(request.user).values_list('id', flat=True))
        _budget_ok = (
            (budget.project_id and budget.project_id in _accessible_ids) or
            (budget.direction_id and budget.direction_id == _be_profile.direction_id)
        )
        if not _budget_ok:
            messages.error(request, _("Vous n'avez pas accès à ce budget."))
            return redirect('core:resources')

    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, _("Budget modifié avec succès."))
            return redirect('core:resources')
    else:
        form = BudgetForm(instance=budget, user=request.user)

    budget_label = budget.project.name if budget.project else (budget.direction.code if budget.direction else 'sans affectation')
    return render(request, 'core/budget_form.html', {'form': form, 'budget': budget, 'title': f'Modifier budget {budget_label}'})


@login_required
def budget_delete(request, budget_id):
    """Supprimer un budget"""
    if not request.user.profile.can_manage_budgets():
        messages.error(request, _("Vous n'avez pas les permissions."))
        return redirect('core:resources')
    
    budget = get_object_or_404(Budget, pk=budget_id)

    _bd_profile = request.user.profile
    _bd_is_global = request.user.is_superuser or _bd_profile.is_admin() or _bd_profile.is_directeur_general()
    if not _bd_is_global:
        _accessible_ids = list(get_accessible_projects_qs(request.user).values_list('id', flat=True))
        _budget_ok = (
            (budget.project_id and budget.project_id in _accessible_ids) or
            (budget.direction_id and budget.direction_id == _bd_profile.direction_id)
        )
        if not _budget_ok:
            messages.error(request, _("Vous n'avez pas accès à ce budget."))
            return redirect('core:resources')

    if request.method == 'POST':
        budget.delete()
        messages.success(request, _("Budget supprimé."))
        return redirect('core:resources')

    return render(request, 'core/confirm_delete.html', {'object': budget, 'type': 'budget', 'back_url': 'core:resources'})


# API endpoints pour les graphiques
@login_required
def api_budget_data(request):
    """API pour les données de budget"""
    can_view_budgets = request.user.profile.can_view_budgets()
    can_view_all_budget_directions = request.user.profile.can_view_all_budget_directions()

    if not can_view_budgets:
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    budgets = Budget.objects.select_related('direction')
    if not can_view_all_budget_directions:
        budgets = budgets.filter(direction=request.user.profile.direction)

    budgets = budgets.all()
    data = []
    for budget in budgets:
        data.append({
            'direction': budget.direction.code if budget.direction else '-',
            'allocated': float(budget.allocated),
            'consumed': float(budget.consumed),
        })
    return JsonResponse(data, safe=False)


@login_required
def api_projects_data(request):
    """API pour les données de projets — scopée selon le rôle de l'utilisateur."""
    accessible = get_accessible_projects_qs(request.user)
    directions = Direction.objects.filter(projects__in=accessible).distinct()
    data = [
        {
            'direction': d.code,
            'en_cours': accessible.filter(direction=d, status='en_cours').count(),
            'termine': accessible.filter(direction=d, status='termine').count(),
        }
        for d in directions
    ]
    return JsonResponse(data, safe=False)


@login_required
def milestone_reorder(request, project_id):
    """API pour réordonner les jalons par drag & drop"""
    import json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    project = get_object_or_404(Project, pk=project_id)

    if not request.user.profile.can_edit_project_milestones(project):
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        data = json.loads(request.body)
        order_list = data.get('order', [])
        for index, milestone_id in enumerate(order_list):
            Milestone.objects.filter(pk=milestone_id, project=project).update(order=index + 1)
        push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def sub_milestone_reorder(request, milestone_id):
    """API pour réordonner les sous-étapes par drag & drop"""
    import json
    from .models import SubMilestone

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    milestone = get_object_or_404(Milestone.objects.select_related('project'), pk=milestone_id)
    project = milestone.project

    if not request.user.profile.can_edit_project_milestones(project):
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        data = json.loads(request.body)
        order_list = data.get('order', [])
        for index, sub_id in enumerate(order_list):
            SubMilestone.objects.filter(pk=sub_id, milestone=milestone).update(order=index + 1)
        push_section_refresh(project.id, 'jalons-view-list', request.user.get_full_name() or request.user.username)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_project_update_status(request, project_id):
    """API pour mettre à jour le statut d'un projet (Kanban drag & drop)"""
    import json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    project = get_object_or_404(Project, pk=project_id)

    if not request.user.profile.can_edit_project(project):
        return JsonResponse({'error': 'Permission refusée'}, status=403)

    try:
        data = json.loads(request.body)
        new_status = data.get('status')
        valid_statuses = [s[0] for s in Project.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({'error': f'Invalid status: {new_status}'}, status=400)
        
        old_status = project.get_status_display()
        project.status = new_status
        project.save(update_fields=['status', 'updated_at'])
        log_project_activity(
            project, 'changement_statut',
            f"Statut changé vers '{project.get_status_display()}'",
            request.user,
        )
        push_project_meta(project)
        return JsonResponse({'success': True, 'status': new_status})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
